from typing import Tuple, List
from base64 import b64encode
from datetime import datetime
from openpyxl.writer.excel import save_virtual_workbook
from copy import deepcopy
from apic.apic_tools import intf_range
from ipaddress import IPv4Address
from ipaddress import AddressValueError
from apic.classes import *
from ipam.utils import BIG, ManagementJob
from data.environments import ACIEnvironment
from smb.SMBConnection import SMBConnection
from smb.base import OperationFailure
from apic.exceptions import *
from githubapi.utils import GithubAPI
from OpenSSL.crypto import FILETYPE_PEM, load_privatekey, sign
from itertools import groupby
from operator import itemgetter
from terraform.utils import Resource
import time
import json
import requests
import random
import os
import openpyxl
import socket
import io
import urllib3
import logging
import string


urllib3.disable_warnings()
urllib3.util.ssl_.DEFAULT_CIPHERS += 'HIGH:!DH:!aNULL'

# TODO: Create a way to check interface usage and status for reclaiming interfaces in the fabric
# TODO: Potentially require the inclusion of the interface profile name

OSP_AUTOMATION_KEY = os.getenv('aci_automation_path')
SUBTREE = 'query-target=subtree'
FULL_CONFIG = 'rsp-subtree=full'
CLASS_FILTER = 'target-subtree-class='
CONFIG_ONLY = 'rsp-prop-include=config-only'
CO = 'rsp-prop-include=config-only'
FCCO = f'{FULL_CONFIG}&{CO}'

SPINE_EVENS = 'Spine-Evens'
SPINE_ODDS = 'Spine-Odds'
ODDS = 'Odds'
EVENS = 'Evens'
OOB = 'OOB'
STAGING = 'Staging'
DRT_TENANT = 'tn-DRTEST'
DRT_VRF = 'vrf-drtest'

EPG_DN_SEARCH = re.compile(r'uni/tn-([^/\]]+)/ap-([^/\]]+)/epg-([^/\]]+)')
BD_DN_SEARCH = re.compile(r'uni/tn-([^/]+)/BD-([^/\]]+)')
AEP_DN_SEARCH = re.compile(r'uni/infra/attentp-([^/\]]+)')
MAC_SEARCH = re.compile(r'([a-f0-9]:){5}[a-f0-9]{2}', flags=re.IGNORECASE)
APIC_MAC_MATCH = re.compile(r'[a-f0-9]{2}(:[a-f0-9]{2}){5}')


def format_mac_addresses(mac_addresses: list) -> List:
    mac_addresses = [''.join(re.findall(r'[a-f0-9]+', _)) for _ in mac_addresses]
    mac_addresses = [':'.join([_.group() for _ in re.finditer(r'[a-f0-9]{2}', mac)]) for mac in mac_addresses]

    # Filter MAC addresses to only include proper MAC addresses
    mac_addresses = [_ for _ in mac_addresses if bool(APIC_MAC_MATCH.fullmatch(_))]

    return mac_addresses


def q_wcard(c, a, v) -> str:
    """Generate query string using wildcard match"""
    return 'query-target-filter=wcard(%s.%s,"%s")' % (c, a, v)


def q_eq(c, a, v) -> str:
    """Generate query string using equal match"""
    return 'query-target-filter=eq(%s.%s,"%s")' % (c, a, v)


class ACIJob:
    queue_path = r'pyapis/aci/job_queue'
    bad_file_path = r'pyapis/aci/invalid'
    completed_jobs = r'pyapis/aci/completed_jobs'
    func: str

    def __init__(self):
        self.run_time = time.time()
        self.environment = ''
        self.configs = []

    @classmethod
    def load(cls, data: dict):
        job = cls()
        for key in data:
            job.__setattr__(key, data[key])

        return job

    @classmethod
    def create_aci_job(cls, job_name: str, environment: str, delay_in_seconds: int, configs: list, func: str):
        gh = GithubAPI()
        job = cls()
        job.environment = environment
        job.run_time += delay_in_seconds
        job.configs = configs
        job.func = func
        gh.add_file(file_path=f'{job.queue_path}/{job_name}.json', message='ACIJob Creation',
                    content=json.dumps(job.__dict__))


class APIC:
    PLACEHOLDERS = 'aep-Placeholders'

    def __init__(self, env: str=None, ip: str=None, username: str=None, password: str=None,
                 use_key: str=OSP_AUTOMATION_KEY):
        if env:
            self.env = ACIEnvironment(env)
            self.ip = socket.gethostbyname(self.env.DNSName)
        elif ip:
            # Check for valid ip
            _ = IPv4Address(ip)
            self.env = None
            self.ip = ip
        else:
            raise ValueError('Either env or ip must be supplied')

        self.session = None
        self.url = f'https://{self.ip}'
        if use_key:
            self.pkey = load_privatekey(FILETYPE_PEM, open(use_key).read())
        else:
            self.pkey = None
            self.login(username=username, password=password)

        self.version = GenericClass.load(json.loads(self.get('/api/class/firmwareCtrlrFwP.json').text)['imdata'][0])
        self.version = self.version.attributes.version.replace('apic-', '')

        if self.env:
            init_leaf = APICObject.load(self.collect_nodes(node_id=self.env.InitLeaf))
            self.leaf_version = init_leaf.attributes.version.replace('n9000-1', '')
        else:
            self.leaf_version = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.pkey:
            pass
        else:
            self.logout()
            self.session.close()

    def __str__(self):
        if self.env:
            return self.env.Name
        else:
            return self.ip

    def login(self, username: str=None, password: str=None):
        self.session = requests.Session()
        self.session.verify = False

        self.url = f'https://{self.ip}'
        creds = {
            'aaaUser': {
                'attributes': {
                    'name': username,
                    'pwd': password
                }
            }
        }
        resp = self.session.post(f'{self.url}/api/aaaLogin.json', json=creds)
        return resp

    def logout(self, username: str=None, password: str=None):
        creds = {
            'aaaUser': {
                'attributes': {
                    'name': (username if username and password else os.getenv('netmgmtuser')),
                    'pwd': (password if username and password else os.getenv('netmgmtpass'))
                }
            }
        }

        self.session.post(f'{self.url}/api/aaaLogout.json', json=creds)
        self.session.close()

    def _sign_request(self, request_string: bytes):
        digest = sign(self.pkey, request_string, digest='sha256')
        signature = b64encode(digest).decode()

        cookie = f'APIC-Request-Signature={signature}; ' \
                 f'APIC-Certificate-Algorithm=v1.0; ' \
                 f'APIC-Certificate-Fingerprint=fingerprint; ' \
                 f'APIC-Certificate-DN=uni/userext/user-automation/usercert-aci_automation'

        return cookie

    def get(self, request: str) -> requests.Response:
        if self.pkey:
            request_string = f'GET{request}'.encode()

            cookie = self._sign_request(request_string)

            return requests.get(f'{self.url}{request}', headers={'Cookie': cookie}, verify=False)
        else:
            return self.session.get(f'{self.url}{request}')

    def post(self, configuration: dict, uri: str='/api/mo/uni.json') -> requests.Response:
        if self.pkey:
            request_string = f'POST{uri}{json.dumps(configuration)}'.encode()

            cookie = self._sign_request(request_string)

            return requests.post(f'{self.url}{uri}', headers={'Cookie': cookie}, json=configuration, verify=False)
        else:
            r = self.session.post(f'{self.url}{uri}', json=configuration)
            if r.status_code == 403:
                print('Session Timed Out: Login Again...')

            return r

    def collect(self, **kwargs) -> list:
        if len(kwargs) > 1:
            raise SyntaxError('This method accepts only one kwarg')

        for key, value in kwargs.items():
            r = self.get(f'/api/class/{key}.json?rsp-subtree=full&rsp-prop-include=config-only').json()

            subjects = r['imdata']

            if value != '':
                try:
                    subject = next(x for x in subjects if x[key]['attributes']['name'] == value)
                except StopIteration:
                    return []
                return [subject]
            else:
                return subjects

    @staticmethod
    def new_bd(name: str, vrf: str, l2: bool=False):
        return {
            'fvBD': {
                'attributes': {
                    'arpFlood': ('no' if l2 is False else 'yes'),
                    'ipLearning': ('yes' if l2 is False else 'no'),
                    'limitIpLearnToSubnets': 'yes',
                    'multiDstPktAct': 'bd-flood',
                    'name': (f'bd-{name}' if not name.startswith('bd-') else name),
                    'type': 'regular',
                    'unicastRoute': ('yes' if l2 is False else 'no'),
                    'unkMacUcastAct': ('proxy' if l2 is False else 'flood'),
                    'unkMcastAct': 'flood'
                },
                'children': [{
                    'fvRsCtx': {
                        'attributes': {
                            'tnFvCtxName': (vrf if vrf.startswith('vrf-') else f'vrf-{vrf}')
                        }
                    }
                }
                ]
            }
        }

    def collect_bds(self, tn: str='', bd: str=''):
        """
        :rtype: dict: if bd is not '': list: if bd is ''
        :param tn:
        :param bd:
        :return:

        Usage:
            APIC.collect_bds(tn: str='', bd: str='')

        Returns:
            List of all BDs that match search criteria.  Dict if BD is specified.
        """

        # Get Bridge Domains from Target Environment
        r = self.get(f'/api/class/fvBD.json?rsp-subtree=full&rsp-prop-include=config-only')
        r = json.loads(r.text)

        bds = r['imdata']

        if tn != '':
            for x in bds[:]:
                for key in x:
                    if f'tn-{tn}' != x[key]['attributes']['dn'].split('/')[1]:
                        bds.remove(x)

        if bd != '':
            try:
                bd = next(x for x in bds if x['fvBD']['attributes']['name'] == bd)
            except StopIteration:
                return 'Bridge Domain does not exist'
            return bd
        else:
            return bds

    def bd_exists(self, tn: str='', bd_name: str=''):
        bd = self.get(f'/api/mo/uni/tn-{tn}/BD-{bd_name}.json?rsp-prop-include=config-only').json()

        if bd['totalCount'] == '1':
            return bd['imdata'][0]
        else:
            return False

    def collect_vrfs(self, tn: str='', vrf: str=''):
        """
        :rtype: list: if vrf is not '': dict if vrf == ''
        :param tn: str
        :param vrf: str
        :return:
        """
        r = self.get(f'/api/class/fvCtx.json?rsp-prop-include=config-only&'
                     f'query-target-filter=wcard(fvCtx.dn,"{tn}")')
        r = json.loads(r.text)
        vrfs = r['imdata']

        if tn != '':
            for x in vrfs[:]:
                for key in x:
                    if f'tn-{tn}' != x[key]['attributes']['dn'].split('/')[1]:
                        vrfs.remove(x)

        if vrf != '':
            try:
                vrf = next(x for x in vrfs if x['fvCtx']['attributes']['name'] == vrf)
            except StopIteration:
                return 'VRF does not exist'
            return vrf
        else:
            return vrfs

    def collect_aps(self, tn: str='', ap: str=''):
        """
        :rtype: dict: if ap is not '': list: if ap is ''
        :param tn:
        :param ap:
        :return:
        """
        r = self.get(f'/api/class/fvAp.json?rsp-subtree=full&rsp-prop-include=config-only')
        r = json.loads(r.text)
        aps = r['imdata']

        if tn != '':
            for x in aps[:]:
                for key in x:
                    if f'tn-{tn}' != x[key]['attributes']['dn'].split('/')[1]:
                        aps.remove(x)

        if ap != '':
            try:
                ap = next(x for x in aps if x['fvAp']['attributes']['name'] == ap)
            except StopIteration:
                return 'Application Profile does not exist'
            return ap
        else:
            return aps

    def collect_tf_aps(self) -> list:
        r = [APICObject.load(_) for _ in self.get(f'/api/class/fvAp.json').json()['imdata']]
        # Filter application profiles to include only production tenants
        r = [_ for _ in r if re.search(rf'/tn-({self.env.Tenant}|{self.env.ADMZTenant})/', _.attributes.dn)]
        r = [_.attributes.annotation.tfref for _ in r if _.attributes.annotation.__dict__.get('tfref')]
        r.sort()
        return r

    def collect_tags(self, tag: str):
        r = self.get(f'/api/class/tagInst.json?query-target-filter=eq(tagInst.name,"{tag}")')
        r = json.loads(r.text)
        tags = r['imdata']

        return tags

    def collect_subnets(self, ip: str=''):
        """
        :rtype: dict: if ip is not '': list: if ip is ''
        :param ip:
        :return:
        """
        bd_subnets = self.get('/api/class/fvSubnet.json?rsp-subtree=full&rsp-prop-include=config-only').json()['imdata']
        ext_subnets = self.get('/api/class/l3extSubnet.json?rsp-prop-include=config-only&query-target-filter='
                               'ne(l3extSubnet.ip,"0.0.0.0/0")').json()['imdata']
        intf_subnets = self.get('/api/class/l3extRsPathL3OutAtt.json?query-target-filter=ne(l3extRsPathL3OutAtt.addr,'
                                '"0.0.0.0")&rsp-prop-include=config-only').json()['imdata']

        for subnet in ext_subnets[:]:
            if 'import-security' not in subnet['l3extSubnet']['attributes']['scope']:
                ext_subnets.remove(subnet)

        bd_subnets = [GenericClass.load(s) for s in bd_subnets]
        ext_subnets = [GenericClass.load(s) for s in ext_subnets]
        intf_subnets = [GenericClass.load(s) for s in intf_subnets]

        networks = []
        for subnet in intf_subnets:
            networks.append([IPv4Network(subnet.attributes.addr, strict=False), subnet])
        for subnet in ext_subnets:
            if 'import-security' in subnet.attributes.scope:
                networks.append([IPv4Network(subnet.attributes.ip, strict=False), subnet])
        for subnet in bd_subnets:
            networks.append([IPv4Network(subnet.attributes.ip, strict=False), subnet])

        if not ip == '':
            matching_networks = []

            ip = IPv4Network(ip)
            for network in networks:
                if network[0].overlaps(ip):
                    matching_networks.append(network)

            matching_networks.sort(key=lambda x: network[0].prefixlen)

            # Returns the smallest network matching the provided IP
            if matching_networks:
                return matching_networks[-1][1].json()
            else:
                return matching_networks
        else:
            return [network[1].json() for network in networks]

    def collect_epgs(self, tn: str='', epg: str=''):
        """
        :rtype: dict: if epg is not '': list: if epg is ''
        :param tn: str
        :param epg: str
        :return:

        Usage:
            APIC.collect_epgs(tn: str='', epg: str='')

        Returns:
            List of all EPGs that match search criteria.  Dict if EPG is specified.
        """
        r = self.get('/api/class/fvAEPg.json?rsp-subtree=full&rsp-prop-include=config-only')
        r = json.loads(r.text)

        epgs = r['imdata']

        if tn != '':
            for x in epgs[:]:
                for key in x:
                    if f'tn-{tn}' != x[key]['attributes']['dn'].split('/')[1]:
                        epgs.remove(x)

        if epg != '':
            try:
                epg = next(x for x in epgs if x['fvAEPg']['attributes']['name'] == epg)
            except StopIteration:
                return 'EPG does not exist'
            return epg

        return epgs

    def collect_aeps(self, aep: str=''):
        result = self.collect(infraAttEntityP=aep)
        return result

    def collect_switch_profiles(self, profile_name: str=''):
        result = self.collect(infraNodeP=profile_name)
        return result

    def collect_pods(self, pod: str=''):
        result = self.get(f'/api/class/fabricPod.json?query-target-filter=wcard(fabricPod.id,"{pod}")').json()['imdata']
        result = sorted(result, key=lambda x: x['fabricPod']['attributes']['id'])
        return result

    def collect_encaps(self, encap=None):
        if_conn = self.get('/api/class/fvIfConn.json')
        if_conn = json.loads(if_conn.text)
        if_conn = if_conn['imdata']

        func_to_epg = self.get('/api/class/infraRsFuncToEpg.json')
        func_to_epg = json.loads(func_to_epg.text)
        func_to_epg = func_to_epg['imdata']

        # Strip dn to EPG
        for x in if_conn:
            x['fvIfConn']['attributes']['dn'] = x['fvIfConn']['attributes']['dn'][
                x['fvIfConn']['attributes']['dn'].index('-[')+2: x['fvIfConn']['attributes']['dn'].index(']/node')]

        for x in func_to_epg:
            x['infraRsFuncToEpg']['attributes']['tDn'] = x['infraRsFuncToEpg']['attributes']['tDn'].replace('uni/', '')

        # Create dictionary using EPGs as keys
        data = {x['fvIfConn']['attributes']['dn']: [] for x in if_conn}

        for x in func_to_epg:
            data[x['infraRsFuncToEpg']['attributes']['tDn']] = []

        # Append encapsulations found to be associated with each EPG
        for key in data:
            for ep in if_conn:
                if ep['fvIfConn']['attributes']['dn'] == key:
                    if ep['fvIfConn']['attributes']['encap'] not in data[key]:
                        data[key].append(ep['fvIfConn']['attributes']['encap'])
            for binding in func_to_epg:
                if binding['infraRsFuncToEpg']['attributes']['tDn'] == key:
                    if binding['infraRsFuncToEpg']['attributes']['encap'] not in data[key]:
                        data[key].append(binding['infraRsFuncToEpg']['attributes']['encap'])

        if encap is not None:
            if isinstance(encap, str):
                if encap.isnumeric():
                    encap = f'vlan-{encap}'
            elif isinstance(encap, int):
                encap = f'vlan-{encap}'
            else:
                raise TypeError('VLAN must be str or int')

            try:
                return next({x: data[x]} for x in data if encap in data[x])
            except StopIteration:
                return 'Encapsulation Is Available'
        else:
            return data

    def collect_static_paths(self, switch: int=None, port: int=None, str_search: str=None):
        sps = self.get('/api/class/fvRsPathAtt.json').json()['imdata']
        if switch is not None and port is not None:
            for sp in sps[:]:
                _sp = GenericClass.load(sp)
                if 'protpath' in _sp.attributes.tDn:
                    sps.remove(sp)
                    continue
                if not switch == int(re.search(r'\d+', _sp.attributes.tDn.split('/')[2]).group()):
                    sps.remove(sp)
                    continue
                elif not port == int(re.search(r'\d+', _sp.attributes.tDn.split('/')[4]).group()):
                    sps.remove(sp)
                    continue
        elif str_search is not None:
            for sp in sps[:]:
                _sp = GenericClass.load(sp)
                if str_search not in _sp.attributes.tDn:
                    sps.remove(sp)
        return sps

    def collect_tdn(self, **kwargs):
        collection = {}

        for arg in kwargs:
            r = self.get('/api/class/{x}.json?query-target-filter=wcard({x}.tDn,"{y}")&rsp-prop-include=config-only'.
                         format(x=arg, y=kwargs[arg]))
            r = json.loads(r.text)
            obj = r['imdata']
            if len(r) > 0:
                collection['{}:{}'.format(arg, kwargs[arg])] = obj
            else:
                continue

        return collection

    def collect_inventory(self):
        r = self.get('/api/class/topSystem.json')
        r = json.loads(r.text)
        inventory = r['imdata']

        inventory.sort(key=lambda x: int(x['topSystem']['attributes']['id']))

        for inv in inventory:
            for attribute in [
                    'bootstrapState'
                    'childAction'
                    'configIssues'
                    'currentTime'
                    'etepAddr'
                    'fabricMAC'
                    'id'
                    'inbMgmtAddr'
                    'inbMgmtAddr6'
                    'inbMgmtAddr6Mask'
                    'inbMgmtAddrMask'
                    'inbMgmtGateway'
                    'inbMgmtGateway6'
                    'lcOwn'
                    'modTs'
                    'mode'
                    'monPolDn'
                    'nameAlias'
                    'nodeType'
                    'oobMgmtAddr6'
                    'oobMgmtAddr6Mask'
                    'oobMgmtGateway6'
                    'remoteNetworkId'
                    'remoteNode'
                    'siteId'
                    'state'
                    'status'
                    'tepPool'
                    'unicastXrEpLearnDisable']:
                try:
                    del inv['topSystem']['attributes'][attribute]
                except KeyError:
                    pass

        return inventory

    def collect_name(self, **kwargs):
        collection = {}

        for arg in kwargs:
            r = self.get('/api/class/{x}.json?query-target-filter=wcard({x}.name,"{y}")'.format(
                x=arg, y=kwargs[arg]))
            r = json.loads(r.text)

            obj = r['imdata']
            if len(r) > 0:
                collection['{}:{}'.format(arg, kwargs[arg])] = obj
            else:
                continue

        return collection

    def collect_eps(self, epg: str=''):
        if epg == '':
            r = self.get('/api/class/fvCEp.json')
            r = json.loads(r.text)

            eps = r['imdata']
            return eps
        else:
            r = self.get(f'{self.url}/api/class/fvCEp.json?query-target-filter=wcard(fvCEp.dn,"{epg}")')
            r = json.loads(r.text)
            r = r['imdata']
            return r

    def collect_eps_by_subnet(self, cidr: str) -> List[GenericClass]:
        cidr = IPv4Network(cidr, strict=False)

        eps = [APICObject.load(_) for _ in self.get('/api/class/fvCEp.json').json()['imdata']]
        eps = [_ for _ in eps if _.attributes.ip if cidr.overlaps(IPv4Network(_.attributes.ip))]

        return eps

    def collect_nodes(self, node_id=None):
        r = self.get('/api/class/fabricNode.json')
        r = json.loads(r.text)

        nodes = r['imdata']

        if node_id is not None:
            try:
                node = next(x for x in nodes if x['fabricNode']['attributes']['id'] == f'{node_id}')
            except StopIteration:
                raise LookupError('Node does not exist')
            return node
        else:
            return nodes

    def collect_interface_profiles(self, name: str=None):
        profiles = self.get('/api/class/infraAccPortP.json').json()['imdata']

        if name:
            profile = next(p for p in profiles if p['infraAccPortP']['attributes']['name'] == name)
            return profile
        else:
            return profiles

    def collect_teps(self):
        r = self.get('/api/class/dhcpClient.json').json()['imdata']
        teps = []

        for _ in r:
            tep = GenericClass.load(_)
            if tep.attributes.nodeRole == 'vip':
                tep = {
                    'name': tep.attributes.id,
                    'address': tep.attributes.ip,
                    'pod': tep.attributes.podId
                }
                if tep not in teps:
                    teps.append(tep)
            elif tep.attributes.nodeRole in ['leaf', 'spine']:
                tep = {
                    'name': tep.attributes.name,
                    'address': tep.attributes.ip,
                    'pod': tep.attributes.podId
                }
                if tep not in teps:
                    teps.append(tep)
            elif tep.attributes.nodeRole == 'protection-chain':
                tep = {
                    'name': tep.attributes.id,
                    'address': tep.attributes.ip,
                    'pod': tep.attributes.podId
                }
                if tep not in teps:
                    teps.append(tep)
            else:
                pass

        return teps

    def collect_snmp_clients(self):
        clients = [GenericClass.load(o) for o in self.get('/api/class/snmpClientP.json').json()['imdata']]
        return clients

    @classmethod
    def w_collect_snmp_clients(cls):
        envs = json.load(open('data/ACIEnvironments.json'))

        clients = []

        for env in envs['Environments']:
            with cls(env=env['Name']) as apic:
                cs = apic.collect_snmp_clients()

            cs = [(c.attributes.name, c.attributes.addr) for c in cs]

            clients += cs

        clients = list(set(clients))

        return 200, clients

    def class_dn_search(self, object_class: str, filter_string: str, config_only=True, get_children: bool=False):
        """
        :rtype: list if len(r['imdata']) > 1: dict if len(r['imdata']) == 1: False if len(r['imdata']) == 0
        :param object_class:
        :param filter_string:
        :param config_only: True
        :param get_children: False
        :return:

        Returns any class objects that match the dn_filter seach string.  To match multiple criteria, separate search
        with asterisk.
        """
        search = []
        filter_string = filter_string.split('*')

        for criteria in filter_string:
            search.append(f'wcard({object_class}.dn,"{criteria}")')

        query = f'/api/class/{object_class}.json?query-target-filter=and({",".join(search)})' \
                f'{("&rsp-prop-include=config-only" if config_only is True else "")}' \
                f'{("" if get_children is False else "&rsp-subtree=full")}'

        r = self.get(query)
        r = json.loads(r.text)

        if len(r['imdata']) > 1:
            return r['imdata']
        elif len(r['imdata']) == 1:
            return r['imdata'][0]
        else:
            return False

    def class_attribute_search(self, object_class: str, attribute: str, value: str, config_only: bool=True,
                               get_children: bool=False):
        """
        :rtype: list if len(r['imdata']) > 1: dict if len(r['imdata'] == 1: False

        :param object_class: str
        :param attribute: str
        :param value: str
        :param config_only: bool=True
        :param get_children: bool=False
        :return:
        """

        search = []
        value = value.split('*')

        for criteria in value:
            search.append(f'wcard({object_class}.{attribute},"{criteria}")')

        query = f'/api/class/{object_class}.json?query-target-filter=and({",".join(search)})' \
                f'{("&rsp-prop-include=config-only" if config_only is True else "")}' \
                f'{("" if get_children is False else "&rsp-subtree=full")}'

        r = self.get(query)
        r = json.loads(r.text)

        if len(r['imdata']) > 1:
            return r['imdata']
        elif len(r['imdata']) == 1:
            return r['imdata'][0]
        else:
            return False

    def dn_exists(self, get: bool=False, **kwargs):
        def complete(exists, obj):
            if 'error' in obj.keys():
                raise NameError('Invalid APIC Class name:', obj['error']['attributes']['text'])

            if get is True:
                return obj
            else:
                return exists

        if len(kwargs) == 1:
            for arg in kwargs:
                if '*' in kwargs[arg]:
                    search = []
                    kwargs[arg] = kwargs[arg].split('*')

                    for criteria in kwargs[arg]:
                        search.append(f'wcard({arg}.dn,"{criteria}")')

                    query = '?query-target-filter=and({})'.format(','.join(search))

                    r = self.get(f'/api/class/{arg}.json{query}')
                    r = json.loads(r.text)
                    if int(r['totalCount']) == 1:
                        result = True
                        return complete(result, r['imdata'][0])
                    else:
                        return False
                else:
                    r = self.get('/api/class/{x}.json?query-target-filter=wcard({x}.dn,"{y}")'.format(
                        x=arg, y=kwargs[arg]
                    ))
                    r = json.loads(r.text)
                    if int(r['totalCount']) == 1:
                        result = True
                        return complete(result, r['imdata'][0])
                    else:
                        return False
        else:
            return 'This module only accepts one keyword argument'

    def exists(self, **kwargs):
        if len(kwargs) == 1:
            for arg in kwargs:
                r = self.get('/api/class/{x}.json?query-target-filter=eq({x}.name,"{y}")'.format(
                    x=arg, y=kwargs[arg]
                ))
                r = json.loads(r.text)
                if int(r['totalCount']) == 1 and 'error' not in r['imdata'][0].keys():
                    return True
                else:
                    return False
        else:
            return 'This module accepts only one keyword argument'

    def create_aep(self, aep_name: str) -> Tuple[int, dict]:
        # Check to see if AEP name exists
        resp = self.get('/api/mo/uni/infra/attentp-%s.json?rsp-subtree=full' % aep_name)

        if not resp.json()['imdata']:
            aep = AEP()
            aep.name = aep_name
            aep.use_domain(self.env.PhysicalDomain)
            aep.children.append(InfraGeneric())

            self.post(aep.json())

        return resp.status_code, resp.json()

    def attach_random_epgs(self, aep_name: str, count: int) -> None:
        rs_epgs = self.get('/api/class/infraRsFuncToEpg.json?%s' % CONFIG_ONLY).json()['imdata']
        rs_epgs = [GenericClass.load(_) for _ in rs_epgs]

        # Get some random EPGs
        attachments = []

        for i in range(count):
            attachments.append(rs_epgs[random.randint(0, len(rs_epgs))])

        # Modify random EPG relationships to the specified AEP
        for a in attachments:
            old_name = re.search(r'attentp-([^/]+)', a.attributes.dn).group()
            a.attributes.dn = a.attributes.dn.replace(old_name, 'attentp-%s' % aep_name)
            a.create()
            print(a.json())
            resp = self.post(a.json())
            print(resp.status_code, resp.json())

    def get_dr_vlans(self):
        func_to_epg = self.class_dn_search(object_class='infraRsFuncToEpg', filter_string='', config_only=True)

        if isinstance(func_to_epg, dict):
            func_to_epg = [func_to_epg]

        response = {'Assignments': []}

        for assignment in func_to_epg:
            func = InfraRsFuncToEpg.load(assignment)
            if 'tn-HCA' in func.attributes.tDn or 'tn-ADMZ' in func.attributes.tDn \
                    or 'tn-common' in func.attributes.tDn or 'tn-mgmt' in func.attributes.tDn \
                    or 'tn-infra' in func.attributes.tDn or 'tn-HCADR' in func.attributes.tDn:
                pass
            else:
                a = {
                        'dr_environment': self.env.Name,
                        'prod_environment': func.attributes.tDn.split('/')[1].replace('tn-tn-', ''),
                        'epg': func.attributes.tDn.split('/')[-1][4:],
                        'vlan': re.search(r'\d+', func.attributes.encap).group()
                    }
                if a not in response['Assignments']:
                    response['Assignments'].append(a)

        return response
    
    def get_vlan_data(self, vlan=None, epg=None, aep=None, dn=None):
        """
        :rtype: dict if vlan is not None: list
        :param vlan:
        :param epg:
        :param aep:
        :param dn:
        :return:
        """
        if_conn = [APICObject.load(_) for _ in self.get('/api/class/fvIfConn.json').json()['imdata']]
        func_to_epg = [APICObject.load(_) for _ in self.get('/api/class/infraRsFuncToEpg.json').json()['imdata']]

        if_conn = [_ for _ in if_conn if re.search(r'vlan-\d+', _.attributes.encap)]

        for record in if_conn:
            # record.encap = int(record['fvIfConn']['attributes']['encap'].replace('vlan-', ''))
            record.encap = int(re.search(r'\d+', record.attributes.encap).group())
            # record['fvIfConn']['attributes']['object'] = \
            #     re.search('tn[-A-Za-z_/0-9]+', record['fvIfConn']['attributes']['dn']).group()
            # print(record.attributes.dn)
            record.object = re.search(r'\[(uni/tn-[^]]+)', record.attributes.dn)
            record.aep = re.search(r'attEntitypathatt-\[([^]]+)', record.attributes.dn)

        for record in func_to_epg:
            # record['infraRsFuncToEpg']['attributes']['encap'] = \
            #     int(record['infraRsFuncToEpg']['attributes']['encap'].replace('vlan-', ''))
            record.encap = int(re.search(r'\d+', record.attributes.encap).group())
            # record['infraRsFuncToEpg']['attributes']['object'] = \
            #     re.search('tn.*', record['infraRsFuncToEpg']['attributes']['tDn']).group()
            record.object = re.search(r'\[(uni/tn-[^]]+)', record.attributes.dn)
            record.aep = AEP_DN_SEARCH.search(record.attributes.dn)

        vlan_dict = {each.encap: {'Consumers': set(), 'AEPs': set()} for each in if_conn + func_to_epg}

        for record in if_conn + func_to_epg:
            if record.object:
                vlan_dict[record.encap]['Consumers'].add(record.object.group(1))
            if record.aep:
                vlan_dict[record.encap]['AEPs'].add(record.aep.group(1))

        for key in vlan_dict:
            vlan_dict[key]['Consumers'] = list(vlan_dict[key]['Consumers'])
            vlan_dict[key]['AEPs'] = list(vlan_dict[key]['AEPs'])

        if vlan is not None:
            try:
                return {vlan: vlan_dict[vlan]}
            except KeyError:
                return 'VLAN Is Not Used'
        elif epg is not None:
            try:
                return list(({key: vlan_dict[key]} for key in vlan_dict
                             for entry in vlan_dict[key]['Consumers']
                             if re.search(f'{epg}$', entry)))
            except StopIteration:
                return ['VLAN Not Found for Endpoint Group']
        elif aep is not None:
            try:
                return list(({key: vlan_dict[key]} for key in vlan_dict
                             for entry in vlan_dict[key]['AEPs']
                             if aep in entry))
            except StopIteration:
                return ['No VLANs Not Found on Port Template']
        elif dn is not None:
            return list(({key: vlan_dict[key]} for key in vlan_dict
                         if dn in vlan_dict[key]['Consumers']))
        else:
            return vlan_dict

    def get_nondr_vlans(self):
        func_to_epg = self.class_dn_search(object_class='infraRsFuncToEpg', filter_string='', config_only=True)

        if isinstance(func_to_epg, dict):
            func_to_epg = [func_to_epg]

        response = {'Assignments': []}

        for assignment in func_to_epg:
            func = InfraRsFuncToEpg.load(assignment)
            if 'tn-HCA' in func.attributes.tDn or 'tn-ADMZ' in func.attributes.tDn:
                a = {
                        'tenant': func.attributes.tDn.split('/')[1][3:],
                        'epg': func.attributes.tDn.split('/')[-1][4:],
                        'vlan': re.search(r'\d+', func.attributes.encap).group()
                    }
                if a not in response['Assignments']:
                    response['Assignments'].append(a)
            else:
                pass

        return response

    def get_vlan_data_old(self, vlan=None, epg=None, aep=None):
        """
        :rtype: dict if vlan is not None: list
        :param vlan:
        :param epg:
        :param aep:
        :return:
        """
        if_conn = self.class_dn_search(object_class='fvIfConn', filter_string='vlan-', config_only=False)
        func_to_epg = self.class_dn_search(object_class='infraRsFuncToEpg', filter_string='', config_only=True)

        if isinstance(func_to_epg, dict):
            func_to_epg = [func_to_epg]

        for record in if_conn:
            record['fvIfConn']['attributes']['encap'] = int(
                record['fvIfConn']['attributes']['encap'].replace('vlan-', ''))
            record['fvIfConn']['attributes']['object'] = \
                re.search('tn[-A-Za-z_/0-9]+', record['fvIfConn']['attributes']['dn']).group()
            try:
                record['fvIfConn']['attributes']['aep'] = \
                    re.search(r'attEnt[-A-Za-z_\[\]0-9]+', record['fvIfConn']['attributes']['dn']).group()
            except AttributeError:
                record['fvIfConn']['attributes']['aep'] = 'AEP Not Found'

        for record in func_to_epg:
            record['infraRsFuncToEpg']['attributes']['encap'] = \
                int(record['infraRsFuncToEpg']['attributes']['encap'].replace('vlan-', ''))
            record['infraRsFuncToEpg']['attributes']['object'] = \
                re.search('tn.*', record['infraRsFuncToEpg']['attributes']['tDn']).group()
            record['infraRsFuncToEpg']['attributes']['aep'] = \
                re.search('attentp-[-A-Za-z_0-9]+', record['infraRsFuncToEpg']['attributes']['dn']).group()

        vlan_dict = {each['fvIfConn']['attributes']['encap']: {'Consumers': [], 'AEPs': []} for each in if_conn}
        for func in func_to_epg:
            vlan_dict[func['infraRsFuncToEpg']['attributes']['encap']] = {'Consumers': [], 'AEPs': []}

        for ep in if_conn:
            vlan_dict[ep['fvIfConn']['attributes']['encap']]['Consumers'].append(ep['fvIfConn']['attributes']['object'])
            if ep['fvIfConn']['attributes']['aep'] != 'AEP Not Found':
                vlan_dict[ep['fvIfConn']['attributes']['encap']]['AEPs'].append(
                    (re.search(r'aep-[-_\w]+', ep['fvIfConn']['attributes']['aep']).group()
                     if re.search(r'aep-[-_\w]+', ep['fvIfConn']['attributes']['aep']) is not None
                     else ep['fvIfConn']['attributes']['aep'])
                )

        for func in func_to_epg:
            vlan_dict[func['infraRsFuncToEpg']['attributes']['encap']]['Consumers'].append(
                func['infraRsFuncToEpg']['attributes']['object']
            )
            vlan_dict[func['infraRsFuncToEpg']['attributes']['encap']]['AEPs'].append(
                (re.search(r'aep-[-_\w]+', func['infraRsFuncToEpg']['attributes']['aep']).group()
                 if re.search(r'aep-[-_\w]+', func['infraRsFuncToEpg']['attributes']['aep']) is not None
                 else func['infraRsFuncToEpg']['attributes']['aep'])
            )

        for key in vlan_dict:
            vlan_dict[key]['Consumers'] = list(set(vlan_dict[key]['Consumers']))
            vlan_dict[key]['AEPs'] = list(set(vlan_dict[key]['AEPs']))

        # Convert vlan_dict keys to integers
        for key in dict(vlan_dict):
            vlan_dict[int(key)] = vlan_dict.pop(key)

        if vlan is not None:
            try:
                return {vlan: vlan_dict[vlan]}
            except KeyError:
                return 'VLAN Is Not Used'
        elif epg is not None:
            try:
                return list(({key: vlan_dict[key]} for key in vlan_dict
                             for entry in vlan_dict[key]['Consumers']
                             if re.search(f'{epg}$', entry)))
            except StopIteration:
                return ['VLAN Not Found for Endpoint Group']
        elif aep is not None:
            try:
                return list(({key: vlan_dict[key]} for key in vlan_dict
                             for entry in vlan_dict[key]['AEPs']
                             if aep in entry))
            except StopIteration:
                return ['No VLANs Not Found on Port Template']
        else:
            return vlan_dict

    def get_next_vlan(self):
        """
        :rtype: int
        :return:
        """
        self.clear_vlans_assigned_to_nonexistent_epgs()
        rs_vlp = self.get(f'/api/mo/uni/phys-{self.env.PhysicalDomain}/rsvlanNs.json').json()['imdata'][0]
        encap_blocks = self.get(f'/api/mo/{rs_vlp["infraRsVlanNs"]["attributes"]["tDn"]}.json?'
                                f'query-target=subtree&target-subtree-class=fvnsEncapBlk').json()['imdata']
        encap_blocks = [EncapBlock.load(_) for _ in encap_blocks]
        encap_blocks.sort(key=lambda x: x.attributes.to)
        vlan_range = [
            '{}-{}'.format(re.search(r"\d+$", _.attributes.__getattribute__("from")).group(),
                           re.search(r"\d+$", _.attributes.to).group())
            for _ in encap_blocks
        ]

        vlan_range = intf_range(vlan_range)
        vlan_range = [_ for _ in vlan_range if _ > 2000]
        used_vlans = self.get_vlan_data()
        return next(x for x in vlan_range if x not in used_vlans)

    def get_next_leaf_pair(self, pod: int=1):
        pod = int(pod)

        r = self.get(f'/api/mo/topology/pod-{pod}.json?'
                     f'query-target=subtree&target-subtree-class=fabricNode').json()['imdata']

        if pod == 1:
            pod = 0
        pod_int = pod * 1000
        assert pod <= 9, 'Pod ID exceeds allowed range'

        node_list = [FabricNode.load(x) for x in r]

        node_list = [int(node.attributes.id) for node in node_list
                     if pod_int + self.env.DataLeafSeries < int(node.attributes.id) < pod_int +
                     (self.env.DataLeafSeries + 99)]

        node_list.sort()

        if not node_list:
            return pod_int + (self.env.DataLeafSeries + 1), pod_int + (self.env.DataLeafSeries + 2)

        # Check for unused, mid-range node IDs
        logical_list = set(range(node_list[0], node_list[-1] + 1))

        available_nodes = logical_list.symmetric_difference(set(node_list))

        if available_nodes:
            available_nodes = list(available_nodes)
            available_nodes.sort()

            if available_nodes[:2] == list(range(available_nodes[0], available_nodes[1] + 1)):
                # This means the two node IDs are sequential and are a good match for return
                node_1, node_2 = available_nodes[:2]

            else:
                # Nodes are not sequential
                raise GetNextLeafError(f'An unexpected error has occurred.  {available_nodes[:2]} are not sequential.')
        else:
            # No leaf IDs available between minimum and maximum, assign next available
            node_1 = node_list[-1] + 1
            node_2 = node_1 + 1

        if node_1 >= pod_int + (self.env.DataLeafSeries + 99):
            raise LeafLimitError

        return node_1, node_2

    def get_next_oob_node(self, pod: int=1):
        pod = int(pod)

        r = self.get(f'/api/mo/topology/pod-{pod}.json?'
                     f'query-target=subtree&target-subtree-class=fabricNode').json()['imdata']

        if pod == 1:
            pod = 0
        pod_int = pod * 1000
        assert pod <= 9, 'Pod ID exceeds allowed range'

        node_list = [FabricNode.load(x) for x in r]

        node_list = [int(node.attributes.id) for node in node_list
                     if pod_int + self.env.MgmtLeafSeries < int(node.attributes.id) < pod_int +
                     (self.env.MgmtLeafSeries + 50)]

        # If this is the first OOB leaf, node_list will be empty.  Return 301 as first OOB node ID
        if not node_list:
            return pod_int + (self.env.MgmtLeafSeries + 1)

        node_list.sort()

        # Check to see if there are any mid-range node IDs available
        logical_list = set(range(node_list[0], node_list[-1] + 1))

        available_nodes = logical_list.symmetric_difference(set(node_list))

        if available_nodes:
            # Found an available ID
            available_nodes = list(available_nodes)
            node_id = available_nodes[0]
        else:
            # No mid-range ID available
            node_id = pod_int + node_list[-1] + 1

        if node_id >= pod_int + (self.env.MgmtLeafSeries + 50):
            raise LeafLimitError

        return node_id

    def get_primary_vrf(self, tenant: str) -> Context:
        vrf = self.get(f'/api/mo/uni/tn-{tenant}.json?query-target=subtree&rsp-prop-include=config-only&'
                       f'query-target-filter=wcard(fvCtx.annotation,"primary_vrf:True")').json()['imdata'][0]
        vrf = Context.load(vrf)
        return vrf

    def set_primary_vrf(self, tenant: str, vrf: str):
        ctxs = self.get(f'/api/mo/uni/tn-{tenant}.json?rsp-prop-include=config-only&query-target=subtree&'
                        f'target-subtree-class=fvCtx').json()['imdata']
        ctxs = [Context.load(_) for _ in ctxs]

        for ctx in ctxs:
            if ctx.attributes.name == vrf:
                ctx.attributes.annotation.__setattr__('primary_vrf', True)
            else:
                ctx.attributes.annotation.__setattr__('primary_vrf', False)

            self.post(configuration=ctx.self_json())

    def verify_serial(self, serial: str):
        r = self.get('/api/class/dhcpClient.json')
        r = json.loads(r.text)['imdata']

        for dhcpClient in r:
            client = GenericClass.load(json_data=dhcpClient)
            if client.attributes.id.upper() == serial.upper():
                return True

        return False

    def assign_leaf_to_maint_group(self, maint_group: str, node: str or int):
        node_block = FabricNodeBlock()
        node_block.attributes.dn = f'uni/fabric/maintgrp-{maint_group}/nodeblk-blk{node}-{node}'
        node_block.attributes.name = f'blk{node}-{node}'
        node_block.attributes.from_ = str(node)
        node_block.attributes.to_ = str(node)
        node_block.attributes.status = 'created,modified'

        r = self.post(configuration=node_block.json())

        return node_block, r

    def assign_leaf_to_firmware_group(self, firmware_group: str, node: str or int):
        node_block = FabricNodeBlock()
        node_block.attributes.dn = f'uni/fabric/fwgrp-{firmware_group}/nodeblk-blk{node}-{node}'
        node_block.attributes.name = f'blk{node}-{node}'
        node_block.attributes.from_ = str(node)
        node_block.attributes.to_ = str(node)
        node_block.attributes.status = 'created,modified'

        r = self.post(configuration=node_block.json())

        return node_block, r

    def assign_oob_to_leaf(self, node_id: str or int, node_name: str):
        node = FabricNode.load(self.collect_nodes(node_id))

        oob_network = IPv4Network(self.env.OOBLeafIPRange, strict=False)

        # Reserve IP address in Proteus
        big = BIG()
        address = big.assign_next_ip(network_cidr=oob_network.with_prefixlen, name=node_name)

        address_cidr = f'{address.properties["address"]}/{oob_network.prefixlen}'
        gateway = f'{oob_network.network_address + 1}'

        big.logout()

        oob_address = OOBAddress()
        oob_address.attributes.addr = address_cidr
        oob_address.attributes.gw = gateway
        oob_address.attributes.tDn = node.attributes.dn

        r = self.post(configuration=oob_address.json(), uri=oob_address.post_uri)

        # Schedule Queue for Management in 2 hours
        ManagementJob.create_mgmt_job(job_name=f'{self.env.Name}_{node_id}_Mgmt', delay_in_seconds=7200,
                                      ip=address.properties['address'], dns_template=self.env.LeafDNSTemplate)

        return oob_address, r

    def register_leaf(self, rack: str, serial: str, node_id: str or int):
        if self.dn_exists(fabricNode=f'node-{node_id}'):
            return 400, ['Node ID selected already exists']

        if self.verify_serial(serial=serial) is False:
            return 400, ['Serial number was not found on fabric']

        if not re.match(r'\w\d\d', rack):
            return 400, ['Invalid Rack number format']

        leaf_name = f'{self.env.Name.upper()}-{rack.upper()}-LF-{node_id}'

        node_policy = NodeIdentityPolicy()
        node_policy.attributes.serial = serial.upper()
        node_policy.attributes.name = leaf_name
        node_policy.attributes.nodeId = str(node_id)

        r = self.post(configuration=node_policy.json(), uri=node_policy.post_uri)

        return node_policy, r

    def add_leaf_to_oob_profile(self, node: str or int):
        switch_profile = SwitchProfile()
        switch_profile.attributes.name = self.env.OOBLeafProfile
        switch_profile.attributes.status = 'created,modified'

        selector = LeafSelector()
        selector.attributes.name = f'LF-{node}'
        selector.attributes.status = 'created'

        switch_profile.children.append(selector)

        block = InfraNodeBlock()
        block.attributes.name = f'{node}-{node}'
        block.attributes.from_ = f'{node}'
        block.attributes.to_ = f'{node}'
        block.attributes.status = 'created'

        selector.children.append(block)

        r = self.post(configuration=switch_profile.json(), uri=switch_profile.post_uri)

        return switch_profile, r

    def create_leaf_pair_profile(self, name: str, nodes: list):
        uni = Uni()

        fabric = Fabric()

        infra = Infra()

        profile = SwitchProfile()
        profile.create_switch_profile(name=name, nodes=nodes)

        fabric_prot_pol = FabricProtPol()
        fabric_prot_pol.add_new_vpc_pair(nodes=nodes)

        infra.children.append(profile)
        fabric.children.append(fabric_prot_pol)

        uni.children.append(infra)
        uni.children.append(fabric)

        r = self.post(configuration=uni.json())

        return uni, r

    @staticmethod
    def check_ssh_to_oob(ip):
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(10)
        try:
            s.connect((ip, 22))
            s.shutdown(socket.SHUT_RDWR)
            s.close()
            return True
        except socket.timeout:
            return False

    def snapshot(self, descr: str):
        descr = descr.replace(' ', '-')
        descr = descr + '-' + str(random.randint(100000, 999999))

        snap = {
            'configExportP': {
                'attributes': {
                    'dn': 'uni/fabric/configexp-defaultOneTime',
                    'name': 'defaultOneTime',
                    'snapshot': 'true',
                    'targetDn': '',
                    'adminSt': 'triggered',
                    'rn': 'configexp-defaultOneTime',
                    'status': 'created,modified',
                    'descr': descr
                }
            }
        }

        self.post(configuration=snap)
        r = self.get(f'/api/class/configExportP?query-target-filter=eq(configExportP.descr,"{descr}")')
        r = json.loads(r.text)

        if r['totalCount'] == '1':
            return descr
        else:
            return False

    def post_to_apic(self, post: dict, uri: str='/api/mo/uni.json'):
        r = self.session.post(f'{self.url}{uri}', json=post)

        if r.ok is True:
            return r
        else:
            return False

    def interface_configuration(self, aep_name: str, infra_info: list):
        assert self.exists(infraAttEntityP=aep_name), 'The specified AEP does not exist'

        # Get complete list of Interface Policy Group Names
        acc_pg_list = self.get('/api/class/infraAccPortGrp.json').json()['imdata']
        pc_pg_list = self.get('/api/class/infraAccBndlGrp.json').json()['imdata']
        acc_pg_list = list((GenericClass.load(x).attributes.name for x in acc_pg_list))
        pc_pg_list = list((GenericClass.load(x).attributes.name for x in pc_pg_list))
        pg_list = acc_pg_list + pc_pg_list

        # Compile complete list of switch profiles that will need to be updated and empty the children containers
        switch_profiles = list((SwitchProfile.load(x) for x in self.collect_switch_profiles()))
        for x in switch_profiles:
            x.attributes.status = 'modified'
            x.children = []

        # Start compiling JSON POST data
        infra = Infra()

        infra_funcp = GenericClass('infraFuncP')  # Policy groups will be a child objects of this object
        infra_funcp.attributes.dn = 'uni/infra/funcprof'
        infra_funcp.attributes.status = 'modified'

        infra.children.append(infra_funcp)

        for server in infra_info:
            interfaces = server['interfaces']
            port_channel = server['port_channel']
            lacp = server['lacp']
            pg_prefix = None

            policy_group = InterfacePolicyGroup()
            policy_group.attributes.status = 'created'
            policy_group.use_aep(aep_name=aep_name)
            if port_channel is True:
                policy_group.port_channel(lacp=lacp)

            for switch_profile in server['switch_profiles']:
                nodes = re.findall(r'\d{3}', switch_profile)
                node_count = len(nodes)
                s_profile = next(x for x in switch_profiles if switch_profile == x.attributes.name)

                if port_channel is True and node_count == 2:
                    i_profile_name = f'vpc-{aep_name.replace("aep-", "")}_{"-".join(nodes)}'
                    if i_profile_name not in list((x.attributes.name for x in infra.children
                                                   if type(x) == InterfaceProfile)):
                        i_profile = InterfaceProfile()
                        i_profile.attributes.name = i_profile_name
                        i_profile.attributes.status = 'created,modified'
                    else:
                        i_profile = next(x for x in infra.children if type(x) == InterfaceProfile
                                         if x.attributes.name == i_profile_name)

                    policy_group.attributes.name = f'vpc-{server["infra_name"]}'
                    pg_prefix = 'vpc'
                    policy_group.attributes.lagT = 'node'

                    selector = InterfaceSelector()
                    selector.attributes.name = f'vpc-{server["infra_name"]}'
                    selector.attributes.descr = server['infra_name']
                    selector.attributes.status = 'created,modified'

                    i_profile.children.append(selector)

                    for intfs in interfaces.split(','):
                        ports = intfs.split('-')
                        block = InterfaceBlock()
                        block.attributes.name = f'block{int(ports[0]) + 1}'
                        block.attributes.fromPort = ports[0]
                        block.attributes.toPort = ports[-1]
                        block.attributes.descr = server['infra_name']

                        selector.children.append(block)

                elif port_channel is True and node_count == 1:
                    i_profile_name = f'pc-{aep_name.replace("aep-", "")}_{"-".join(nodes)}'
                    if i_profile_name not in list((x.attributes.name for x in infra.children
                                                   if type(x) == InterfaceProfile)):
                        i_profile = InterfaceProfile()
                        i_profile.attributes.name = i_profile_name
                        i_profile.attributes.status = 'created,modified'
                    else:
                        i_profile = next(x for x in infra.children if type(x) == InterfaceProfile
                                         if x.attributes.name == i_profile_name)

                    pg_prefix = 'pc'
                    policy_group.attributes.lagT = 'link'

                    selector = InterfaceSelector()
                    selector.attributes.name = f'pc-{server["infra_name"]}'
                    selector.attributes.descr = server['infra_name']
                    selector.attributes.status = 'created,modified'

                    i_profile.children.append(selector)

                    for intfs in interfaces.split(','):
                        ports = intfs.split('-')
                        block = InterfaceBlock()
                        block.attributes.name = f'block{int(ports[0]) + 1}'
                        block.attributes.fromPort = ports[0]
                        block.attributes.toPort = ports[-1]
                        block.attributes.descr = server['infra_name']

                        selector.children.append(block)

                elif port_channel is False:
                    i_profile_name = f'acc-{aep_name.replace("aep-", "")}_{"-".join(nodes)}'
                    if i_profile_name not in list((x.attributes.name for x in infra.children
                                                   if type(x) == InterfaceProfile)):
                        i_profile = InterfaceProfile()
                        i_profile.attributes.name = i_profile_name
                        i_profile.attributes.status = 'created,modified'
                    else:
                        i_profile = next(x for x in infra.children if type(x) == InterfaceProfile
                                         if x.attributes.name == i_profile_name)

                    pg_prefix = 'acc'

                    selector = InterfaceSelector()
                    selector.attributes.name = f'acc-{server["infra_name"]}'
                    selector.attributes.descr = server['infra_name']
                    selector.attributes.status = 'created,modified'

                    i_profile.children.append(selector)

                    for intfs in interfaces.split(','):
                        ports = intfs.split('-')
                        block = InterfaceBlock()
                        block.attributes.name = f'block{int(ports[0]) + 1}'
                        block.attributes.fromPort = ports[0]
                        block.attributes.toPort = ports[-1]
                        block.attributes.descr = server['infra_name']

                        selector.children.append(block)

                else:
                    raise Exception('The Request could not be processed')

                attach_i_profile = GenericClass('infraRsAccPortP')
                attach_i_profile.attributes.tDn = f'uni/infra/accportprof-{i_profile.attributes.name}'
                attach_i_profile.attributes.status = 'created,modified'

                if attach_i_profile.attributes.tDn not in list((x.attributes.tDn for x in s_profile.children)):
                    s_profile.children.append(attach_i_profile)

                if i_profile not in infra.children:
                    infra.children.append(i_profile)

                attach_policy_group = GenericClass('infraRsAccBaseGrp')
                if pg_prefix == 'vpc' or pg_prefix == 'pc':
                    tdn_path = 'uni/infra/funcprof/accbundle-'
                else:
                    tdn_path = 'uni/infra/funcprof/accportgrp-'
                attach_policy_group.attributes.tDn = f'{tdn_path}{pg_prefix}-{server["infra_name"]}'
                attach_policy_group.attributes.status = 'created,modified'

                selector.children.append(attach_policy_group)

            policy_group.attributes.name = f'{pg_prefix}-{server["infra_name"]}'
            if policy_group.attributes.name not in pg_list and \
                    policy_group.attributes.name not in \
                    (x.attributes.name for x in infra_funcp.children if type(x) is InterfacePolicyGroup):
                infra_funcp.children.append(policy_group)
        for x in switch_profiles:
            if not x.children == list():
                infra.children.append(x)

        r = self.post(configuration=infra.json(), uri='/api/mo/uni/infra.json')
        # return infra.json()
        return r, infra.json()

    def update_snmp_strings(self):
        new_communities = requests.post('https://pyapis.ocp.app.medcity.net/apis/admin/get_current_snmp_strings',
                                        json={'APIKey': os.getenv('localapikey'), 'Trusted': True}, verify=False).json()

        communities = self.get('/api/class/snmpCommunityP.json?rsp-prop-include=config-only').json()['imdata']

        for community in communities:
            community = GenericClass.load(community)
            community.attributes.status = 'deleted'
            self.post(community.json())

        community = GenericClass.load({'snmpCommunityP': {'attributes': {}}})
        community.attributes.name = new_communities['ro']
        community.attributes.dn = f'uni/fabric/snmppol-default/community-{new_communities["ro"]}'
        community.attributes.descr = 'RO'
        community.attributes.status = 'created'
        self.post(community.json())

        community.attributes.name = new_communities['rw']
        community.attributes.dn = community.attributes.dn.replace(new_communities['ro'], new_communities['rw'])
        community.attributes.descr = 'RW'
        self.post(community.json())
        return None

    def cleanup_interface_policies(self):
        # Collect unused interface policy groups
        policy_groups = [GenericClass.load(group) for group in
                         self.get('/api/class/infraAccPortGrp.json?rsp-prop-include=config-only').json()['imdata']] + \
                        [GenericClass.load(group) for group in
                         self.get('/api/class/infraAccBndlGrp.json?rsp-prop-include=config-only').json()['imdata']]

        used_policy_groups = [GenericClass.load(group) for group in
                              self.get('/api/class/infraRsAccBaseGrp.json?rsp-prop-include=config-only').json()[
                                  'imdata']]
        used_policy_groups = [group.attributes.tDn for group in used_policy_groups]

        unused_policy_groups = [group for group in policy_groups if group.attributes.dn not in used_policy_groups]

        print('Interface policy groups that are not used by any interface profile')
        print([group.attributes.dn for group in unused_policy_groups])

        # Prompt to delete each unused group
        for group in unused_policy_groups:
            answer = input(f'\n\nDelete this Policy Group? (yes/no)   {group.attributes.dn}')
            if answer.upper().startswith('Y'):
                group.attributes.status = 'deleted'
                self.post(group.json())
            else:
                pass

        # Collect unused interface profiles
        profiles = [GenericClass.load(profile) for profile in
                    self.get('/api/class/infraAccPortP.json?rsp-prop-include=config-only').json()['imdata']]

        used_profiles = [GenericClass.load(profile) for profile in
                         self.get('/api/class/infraRsAccPortP.json?rsp-prop-include=config-only').json()['imdata']]
        used_profiles = [profile.attributes.tDn for profile in used_profiles]

        unused_profiles = [profile for profile in profiles if profile.attributes.dn not in used_profiles]

        # Prompt to delete each unused profile
        for profile in unused_profiles:
            answer = input(f'Delete this Interface Profile?  (yes/no)  ')
            if answer.upper().startswith('Y'):
                profile.attributes.status = 'deleted'
                self.post(profile.json())
            else:
                pass

    def bd_cleanup(self, tenant: str):
        assert tenant.startswith('tn-')
        bds = self.get('/api/class/fvBD.json?rsp-prop-include=config-only').json()['imdata']
        bds = [GenericClass.load(bd) for bd in bds]
        bds = [bd for bd in bds if bd.attributes.dn.split('/')[1] == f'tn-{tenant}']

        used_bds = self.get('/api/class/fvRsBd.json').json()['imdata']
        used_bds = [GenericClass.load(bd) for bd in used_bds]
        used_bds = [bd for bd in used_bds if bd.attributes.dn.split('/')[1] == f'tn-{tenant}']

        for bd in bds:
            if bd.attributes.name not in [ubd.attributes.tnFvBDName for ubd in used_bds]:
                answer = input(f'Do you want to delete {bd.attributes.name} from {tenant}?  (y/n) ')

                if answer.lower().startswith('y'):
                    bd.attributes.__setattr__('status', 'deleted')

                    response = self.post(bd.json())
                    print(response.status_code, response.reason)
        return None

    def fault_cleanup(self):
        # Snapshot first
        self.snapshot('fault cleanup')

        # Collect and delete unresolved references
        faults = self.get('/api/class/faultInst.json').json()['imdata']
        faults = [GenericClass.load(fault) for fault in faults
                  if fault['faultInst']['attributes']['cause'] == 'resolution-failed']

        for fault in faults:
            subject = fault.attributes.dn.split('/')
            subject = '/'.join(subject[:-1])

            # Get the object referencing something that does not exist
            response = self.get(f'/api/mo/{subject}.json?rsp-prop-include=config-only')

            # Check that the response was good and that only one object was returned
            if not response.ok or response.json()['totalCount'] != '1':
                continue

            o = GenericClass.load(response.json()['imdata'][0])
            o.attributes.__setattr__('status', 'deleted')

            # Delete the object causing the fault
            self.post(configuration=o.json())
            # print(o.json())
        return None

    def assert_physical_domain_usage_f0467(self, fix: bool=False):
        """This is a matter of ensuring that all EPGs and AEPs have a physical domain association"""
        aep_dom = GenericClass(apic_class='infraRsDomP')
        aep_dom.attributes.tDn = f'uni/phys-{self.env.PhysicalDomain}'
        aep_dom.create()

        epg_dom = GenericClass(apic_class='fvRsDomAtt')
        epg_dom.attributes.tDn = f'uni/phys-{self.env.PhysicalDomain}'
        epg_dom.create()

        aeps = self.collect_aeps()
        aeps = [AEP.load(_) for _ in aeps]
        aeps = [aep for aep in aeps if aep.attributes.name != 'default']
        aeps = [aep for aep in aeps if 'infraRsDomP' not in [_.class_ for _ in aep.children]]
        aeps = [aep.attributes.dn for aep in aeps]

        faults = self.get('/api/class/faultInst.json?query-target-filter=eq(faultInst.code,"F0467")').json()['imdata']
        faults = [GenericClass.load(fault) for fault in faults
                  if fault['faultInst']['attributes']['code'] == 'F0467']

        epgs = set([re.search(r'(tn-.*?/ap-.*?/epg-.*?)]', _.attributes.dn)[1] for _ in faults
                    if re.search(r'(tn-.*?/ap-.*?/epg-.*?)]', _.attributes.dn)])

        if fix:
            for epg in epgs:
                _ = self.post(configuration=epg_dom.json(), uri=f'/api/mo/uni/{epg}.json')
                print(_.json())

            for aep in aeps:
                _ = self.post(configuration=aep_dom.json(), uri=f'/api/mo/{aep}.json')
                print(_.json())

            return None
        else:
            return aeps, epgs

    def clear_vlans_assigned_to_nonexistent_epgs(self) -> None:
        """Deletes VLANs assigned to AEPs for EPGs that do not exist"""
        faults = self.get('/api/class/faultInst.json?query-target-filter=eq(faultInst.code,"F0982")')
        faults = [APICObject.load(_) for _ in faults.json()['imdata']]

        for fault in faults:
            subject = re.search(r'(.*)/fault-F0982', fault.attributes.dn).group(1)
            subject = APICObject.load(self.get(f'/api/mo/{subject}.json?{CO}').json()['imdata'])
            subject.delete()
            self.post(subject.json())

    def reassign_encap(self, old_epg_dn: str, new_epg_dn: str):
        # Ensure that each variable starts with uni/
        if not old_epg_dn.startswith('uni/'):
            old_epg_dn = f'uni/{old_epg_dn}'
        if not new_epg_dn.startswith('uni/'):
            new_epg_dn = f'uni/{new_epg_dn}'

        # Modify aep attachments to use new EPG
        attachments = self.get('/api/class/infraRsFuncToEpg.json?rsp-prop-include=config-only').json()['imdata']
        attachments = [InfraRsFuncToEpg.load(attachment) for attachment in attachments]
        attachments = [attachment for attachment in attachments if old_epg_dn == attachment.attributes.tDn]

        for attachment in attachments:
            # Delete encap with old EPG on it
            attachment.attributes.__setattr__('status', 'deleted')
            self.post(attachment.json())
            # Create encap with new EPG on it
            attachment.attributes.__delattr__('tDn')
            attachment.attributes.__setattr__('status', 'created,modified')
            attachment.attributes.dn = attachment.attributes.dn.replace(old_epg_dn, new_epg_dn)
            self.post(attachment.json())

        return None

    def change_epg_encap(self, epg_dn: str, old_encap: int, new_encap: int):
        # Ensure that each variable starts with uni/
        if not epg_dn.startswith('uni/'):
            epg_dn = f'uni/{epg_dn}'

        # Modify aep attachments to use new EPG
        attachments = self.get('/api/class/infraRsFuncToEpg.json?rsp-prop-include=config-only').json()['imdata']
        attachments = [InfraRsFuncToEpg.load(attachment) for attachment in attachments]
        attachments = [attachment for attachment in attachments if epg_dn == attachment.attributes.tDn]

        response_data = []

        for attachment in attachments:
            # Only change encapsulation if old_encap matches what was pulled from APIC
            if attachment.attributes.encap == f'vlan-{old_encap}':
                # Modify encap
                attachment.attributes.__delattr__('tDn')
                attachment.attributes.__setattr__('status', 'modified')
                attachment.attributes.encap = f'vlan-{(new_encap if new_encap else self.get_next_vlan())}'
                self.post(attachment.json())
                response_data.append(attachment.json())

        return response_data

    def rebrand_aep(self, old_aep_name: str, new_aep_name: str) -> Tuple[int, dict]:
        return_data = {}

        # Get the old AEP
        resp = self.get('/api/mo/uni/infra/attentp-%s.json?%s&%s' % (old_aep_name, FULL_CONFIG, CONFIG_ONLY))

        if resp.json()['imdata']:
            old_aep = AEP.load(resp.json()['imdata'][0])
        else:
            return_data['get_old_aep_resp'] = [resp.status_code, resp.json()]
            return 400, {'message': 'The old AEP name supplied does not exist. Check spelling and case.',
                         'details': return_data}

        # Create a snapshot in case this need to be rolled back
        return_data['snapshot_name'] = self.snapshot('rebrand_aep')

        # Check if the new AEP name already exists
        new_aep = self.get('/api/mo/uni/infra/attentp-%s.json' % new_aep_name).json()['imdata']

        if new_aep:
            new_existed = True
            new_aep = AEP.load(new_aep[0])
        else:
            new_existed = False
            new_aep = AEP()
            new_aep.name = new_aep_name
            new_aep.attributes.dn = 'uni/infra/attentp-%s' % new_aep_name
            new_aep.use_domain(self.env.PhysicalDomain)
            new_aep.children = [_ for _ in old_aep.children if _.class_ == 'infraRsDomP']  # Keep domains from old
            # Create and add gen-default to new AEP
            gen_default = InfraGeneric()
            new_aep.children += [gen_default]

            old_gen_default = next(_ for _ in old_aep.children if _.class_ == 'infraGeneric')  # Get old gen-default
            gen_default.children += [_ for _ in old_gen_default.children if _.class_ == InfraRsFuncToEpg.class_]
            new_aep.create()

            resp = self.post(new_aep.json())

            return_data['new_aep_create_resp'] = [resp.status_code, resp.json()]
        return_data['new_aep_existed'] = new_existed

        # Check for EPG to VLAN ID conflicts between the AEPs if the new AEP already existed
        if new_existed:
            set_1 = self.get(f'/api/mo/{old_aep.attributes.dn}.json'
                             f'?%s&%sinfraRsFuncToEpg' % (SUBTREE, CLASS_FILTER)).json()['imdata']
            set_2 = self.get(f'/api/mo/{new_aep.attributes.dn}.json'
                             f'?%s&%sinfraRsFuncToEpg' % (SUBTREE, CLASS_FILTER)).json()['imdata']

            for rs_epg in [GenericClass.load(_) for _ in set_1]:
                try:
                    r_2 = next(GenericClass.load(_) for _ in set_2
                               if _['infraRsFuncToEpg']['attributes']['tDn'] == rs_epg.attributes.tDn)
                except StopIteration:
                    continue

                if rs_epg.attributes.encap != r_2.attributes.encap:
                    return_data['conflict'] = {
                        'old_epg_def': {
                            'vlan': rs_epg.attributes.encap,
                            'dn': rs_epg.attributes.dn
                        },
                        'new_epg_def': {
                            'vlan': r_2.attributes.encap,
                            'dn': r_2.attributes.dn
                        }
                    }
                    return 400, {'message': 'VLAN conflict exists between the two AEPs.  Resolve and try again',
                                 'details': return_data}

            # No conflicts exist, AEPs can be sync'd; Only runs if new AEP already existed
            self.sync_aep_attachments(new_aep_name, old_aep_name)

        # Replace relationships to old AEP with relationships to new AEP
        rels = self.get('/api/class/infraRsAttEntP.json'
                        '?%s&%s' % (q_eq('infraRsAttEntP', 'tDn', old_aep.attributes.dn), CONFIG_ONLY)).json()['imdata']

        rel_replacement = []
        for rel in [GenericClass.load(_) for _ in rels]:
            rel.attributes.tDn = new_aep.attributes.dn
            rel.modify()

            resp = self.post(rel.json())
            rel_replacement += [rel.attributes.dn, resp.status_code, resp.json()]

            print(rel.json())

        return_data['rel_replacement'] = rel_replacement

        # Finally, delete the old AEP
        old_aep.attributes.status = 'deleted'
        resp = self.post(old_aep.self_json())

        return_data['old_aep_deletion'] = [resp.status_code, resp.json()]
        print(old_aep.self_json())

        return 200, {'message': 'Old AEP has been replaced by new AEP',
                     'details': return_data}

    def clone_aep(self, aep_name, new_aep_name) -> Tuple[int, dict]:
        # Check if requested AEP exists first
        exists = self.collect_aeps(aep=new_aep_name)
        if exists:
            return 400, {'message': 'The requested aep name already exists'}

        aep = AEP.load(self.collect_aeps(aep_name))
        aep.attributes.__delattr__('dn')
        aep.attributes.name = new_aep_name
        resp = self.post(aep.json(), aep.post_uri)

        return resp.status_code, resp.json()

    def rebrand_epg_bd(self, old_epg_dn: str, new_epg_dn: str, new_bd_name: str=None):
        if not old_epg_dn.startswith('uni/'):
            old_epg_dn = f'uni/{old_epg_dn}'
        if not new_epg_dn.startswith('uni/'):
            new_epg_dn = f'uni/{new_epg_dn}'

        tenant = old_epg_dn.split('/')[1][3:]
        new_tenant = new_epg_dn.split('/')[1][3:]
        epg_name = old_epg_dn.split('/')[-1][4:]

        # Get the primary VRF for the destination tenant.  It may be different that the origin tenant
        ctx = self.get_primary_vrf(tenant=new_tenant)

        # Determine AP name and whether the new AP needs to be created.  Create if needed
        new_ap_name = new_epg_dn.split('/')[2][3:]

        if not new_ap_name.startswith('ap-'):
            new_ap_name = f'ap-{new_ap_name}'

        aps = [AP.load(ap) for ap in self.collect_aps(tn=new_tenant)]
        if new_ap_name not in [ap.attributes.name for ap in aps]:
            ap = AP()
            ap.attributes.dn = f'uni/tn-{new_tenant}/ap-{new_ap_name}'
            ap.attributes.__setattr__('status', 'created')
        else:
            ap = None

        # Check to see if the new EPG name already exists.  Should not continue if it does
        epg_usage = self.get(f'/api/mo/{new_epg_dn}.json')

        if not epg_usage.ok:
            return {'message': 'The existence of the new EPG could not be confirmed',
                    'details': epg_usage.json()}

        if int(epg_usage.json()['totalCount']):
            return {'message': 'The new EPG name provided already exists.  Automation aborted',
                    'details': epg_usage.json()}

        # Collect existing EPG
        epg = EPG.load(self.collect_epgs(tn=tenant, epg=epg_name))

        bd_name = next(child.attributes.tnFvBDName for child in epg.children if child.class_ == 'fvRsBd')

        # Collect all BDs from the new tenant.  May or may not be moving to another Tenant
        bds = [BD.load(_) for _ in self.collect_bds(tn=new_tenant)]

        if new_bd_name and bd_name != new_bd_name:

            # Check to ensure the old BD is not used by multiple EPGs
            multi_usage = self.get(f'/api/mo/uni/tn-{tenant}.json?query-target=subtree&target-subtree-class=fvRsBd'
                                   f'&query-target-filter=eq(fvRsBd.tnFvBDName,"{bd_name}")')
            if not multi_usage.ok:
                # Without a valid response regarding bridge domain usage it would be dangerous to proceed
                return {'message': 'Bad response from APIC when attempting to get bridge domain usage',
                        'details': multi_usage.json()}

            if int(multi_usage.json()['totalCount']) > 1:
                # This scenario could result in outages.  Deleting or changing a bridge domain used by multiple EPGs
                # can result in the loss of endpoints.  The automation exits if this condition is met.
                return {'message': 'BD is in use by multiple EPGs. This prevents this automation from running.'}

            # A new BD was specified. Checks will be run to validate the new name against the existing name
            bd = BD.load(self.collect_bds(tn=tenant, bd=bd_name))

            if new_bd_name in [x.attributes.name for x in bds]:
                # This scenario requires the subnets from the old BD to be put on the existing BD and the old BD be
                # deleted
                # No need to change the VRF here.  The BD in the destination tenant already exists
                new_bd = next(_ for _ in bds if _.attributes.name == new_bd_name)
                new_bd.attributes.__setattr__('status', 'modified')
                # This assigns the subnets from the old BD to the new BD without altering the BD configuration
                new_bd.children = [Subnet.load(child.json()) for child in bd.children if child.class_ == 'fvSubnet']
                # Prepare the old BD for deletion
                bd.attributes.__setattr__('status', 'deleted')
            else:
                # This scenario requires the creation of a new BD.  The old BD will be copied, then deleted assuming it
                # only has one usage
                new_bd = BD.load(bd.json())
                new_bd.attributes.__delattr__('name')
                new_bd.tenant = new_tenant
                new_bd.name = new_bd_name
                print(new_bd.json())
                # New BD means we need to use the primary VRF, most likely
                new_bd.use_vrf(name=ctx.attributes.name)
                print(new_bd.json())
                new_bd.attributes.__setattr__('status', 'created')
                # Prepare the old BD for deletion
                bd.attributes.__setattr__('status', 'deleted')
        else:
            bd = None
            new_bd = None

        # Copy existing EPG to new EPG and modify attributes
        new_epg = EPG.load(epg.json())
        new_epg.attributes.__setattr__('status', 'created')
        new_epg.attributes.__delattr__('name')
        new_epg.attributes.dn = new_epg_dn
        if new_bd:
            rs_bd = next(child for child in new_epg.children if child.class_ == 'fvRsBd')
            rs_bd.attributes.tnFvBDName = new_bd_name

        # Stage existing EPG for deletion
        epg.attributes.__setattr__('status', 'deleted')

        # Implement the configs
        if ap:
            ap_result = self.post(ap.json()).reason
            print(ap.json())
        else:
            ap_result = None

        if bd and new_bd:
            old_bd_result = self.post(configuration=bd.json()).reason
            print(bd.json())
            new_bd_result = self.post(configuration=new_bd.json()).reason
            print(new_bd.json())
        else:
            new_bd_result = None
            old_bd_result = None

        old_epg_result = self.post(epg.json()).reason
        new_epg_result = self.post(new_epg.json()).reason

        self.reassign_encap(old_epg_dn=epg.attributes.dn, new_epg_dn=new_epg.attributes.dn)

        return {'ap_result': ap_result, 'new_epg_result': new_epg_result, 'old_epg_result': old_epg_result,
                'old_bd_result': old_bd_result, 'new_bd_result': new_bd_result}

    def sync_aep_attachments(self, aepa: str, aepb: str):
        aepa = AEP.load(self.collect_aeps(aep=aepa))
        aepb = AEP.load(self.collect_aeps(aep=aepb))

        aepa_attachments = next(x.children for x in aepa.children if x.class_ == 'infraGeneric')
        aepb_attachments = next(x.children for x in aepb.children if x.class_ == 'infraGeneric')

        aepa_attachments = [(obj.attributes.tDn, obj.attributes.encap, obj.attributes.mode) for obj in aepa_attachments]
        aepb_attachments = [(obj.attributes.tDn, obj.attributes.encap, obj.attributes.mode) for obj in aepb_attachments]

        difference = set(aepa_attachments).symmetric_difference(set(aepb_attachments))

        for aep in [aepa, aepb]:
            for tdn, encap, mode in difference:
                attach = InfraRsFuncToEpg()
                attach.attributes.dn = f'uni/infra/attentp-{aep.attributes.name}/gen-default/rsfuncToEpg-[{tdn}]'
                attach.attributes.encap = encap
                attach.attributes.mode = mode
                attach.attributes.status = 'created'

                # Each request where the infraRsFuncToEpg already exists under the AEP is designed to fail
                # Expect Response(400)
                self.post(configuration=attach.json())

    @classmethod
    def move_to_dr(cls, epg: str):
        """Class method used only in SEDC environment.  Migrates an EPG that is typically in tn-HCA into a DR testing
        environment.  It then updates all infraRsFuncToEpgs to use the new EPG created in tn-DRTEST."""
        with cls(env='sedc') as apic:
            # Check for and create DR testing AP in tn-DRTEST
            aps = [GenericClass.load(ap) for ap in apic.collect_aps(tn='tn-DRTEST')]
            if 'ap-DRTESTing' not in [ap.attributes.name for ap in aps]:
                ap = AP()
                ap.attributes.dn = 'uni/tn-tn-DRTEST/ap-ap-DRTESTing'
                ap.attributes.__setattr__('status', 'created')

                ap_result = apic.post(ap.json()).reason
                print(ap.json())
            else:
                ap_result = 'Already Exists'

            # Copy EPG
            epg = EPG.load(apic.collect_epgs(tn='tn-HCA', epg=epg))
            or_epg_dn = epg.attributes.dn
            for child in epg.children[:]:
                if child.class_ not in ['fvRsBd', 'fvRsDomAtt']:
                    epg.children.remove(child)
                elif child.attributes.json() == {}:
                    epg.children.remove(child)

            epg.attributes.dn = f'uni/tn-tn-DRTEST/ap-ap-DRTESTing/epg-{epg.attributes.name}'
            epg.attributes.__setattr__('status', 'created,modified')

            in_use_bd = next(child.attributes.tnFvBDName for child in epg.children if child.class_ == 'fvRsBd')

            # Copy Bridge Domain
            bd = BD.load(apic.collect_bds(tn='tn-HCA', bd=in_use_bd))
            bd.attributes.dn = f'uni/tn-tn-DRTEST/BD-{bd.attributes.name}'
            bd.attributes.__setattr__('status', 'created,modified')

            for child in bd.children[:]:
                if child.class_ == 'fvRsBDToOut':
                    child.attributes.tnL3extOutName = 'L3Out-Core-DRTEST'
                elif child.class_ == 'fvRsCtx':
                    child.attributes.tnFvCtxName = 'vrf-drtest'
                elif child.attributes.json() == {}:
                    bd.children.remove(child)

            # Create BD and EPG for DR testing
            bd_result = apic.post(bd.json()).reason
            print(bd.json())
            epg_result = apic.post(epg.json()).reason
            print(epg.json())

            # Modify aep attachments to use copied EPG
            apic.reassign_encap(old_epg_dn=or_epg_dn, new_epg_dn=epg.attributes.dn)

        return {'ap_result': ap_result, 'bd_result': bd_result, 'epg_result': epg_result}

    @classmethod
    def return_to_prod(cls, epg: str):
        with cls(env='sedc') as apic:
            drtest_dn = f'uni/tn-tn-DRTEST/ap-ap-DRTESTing/epg-{epg}'
            epg = EPG.load(apic.collect_epgs(tn='tn-HCA', epg=epg))

            apic.reassign_encap(old_epg_dn=drtest_dn, new_epg_dn=epg.attributes.dn)

        return None

    @classmethod
    def get_bd_by_ip(cls, environment: str, ip: str):
        with cls(env=environment) as apic:
            subnet = apic.collect_subnets(ip=ip)

            if isinstance(subnet, str):
                return 404, {'message': 'No Subnet found matching IP'}
            else:
                subnet = GenericClass.load(subnet)

            if subnet.class_ == 'l3extRsPathL3OutAtt':
                l3out = re.search('out-(.*)', subnet.attributes.dn.split('/')[2])[1]
                intf_type = subnet.attributes.ifInstT
                encap = subnet.attributes.encap
                leaf = re.search('paths-(.*)', subnet.attributes.tDn.split('/')[2])[1]
                subnet = str(IPv4Network(subnet.attributes.addr, strict=False))
                return 200, {'environment': apic.env.Name, 'l3out': l3out, 'intf_type': intf_type, 'encap': encap,
                             'leaf': leaf, 'subnet': subnet}
            elif subnet.class_ == 'l3extSubnet':
                print(subnet.json())
                l3out = re.search('out-(.*)', subnet.attributes.dn.split('/')[2])[1]
                external_epg = re.search('instP-(.*)', subnet.attributes.dn.split('/')[3])[1]
                subnet = str(IPv4Network(subnet.attributes.ip, strict=False))
                return 200, {'environment': apic.env.Name, 'l3out': l3out, 'external_epg': external_epg,
                             'subnet': subnet}
            elif subnet.class_ == 'fvSubnet':
                bridge_domain = re.search('BD-(.*)', subnet.attributes.dn.split('/')[2])[1]
                tenant = re.search('tn-(.*)', subnet.attributes.dn.split('/')[1])[1]
                bd = BD.load(apic.collect_bds(tn=tenant, bd=bridge_domain))
                rsbds = apic.get(f'/api/class/fvRsBd.json?query-target-filter=eq(fvRsBd.tnFvBDName,'
                                 f'"{bd.attributes.name}"').json()['imdata']
                rsbds = [GenericClass.load(rsbd).attributes.dn.strip('/rsbd') for rsbd in rsbds]
                subnets = [str(IPv4Network(s.attributes.ip, strict=False)) for s in bd.children
                           if s.class_ == Subnet.class_]
                subnet = str(IPv4Network(subnet.attributes.ip, strict=False))
                return 200, {'environment': apic.env.Name, 'tenant': tenant, 'bridge_domain': bridge_domain,
                             'subnet': subnet, 'all_bd_subnets': subnets, 'inuse_by': rsbds}

    @classmethod
    def tag_epgs(cls, env: str, epgs: list):
        with cls(env=env) as apic:
            response_data = {}

            for epg in epgs:
                epg_resp = apic.collect_epgs(epg=epg['epg_name'])

                if epg_resp == 'EPG does not exist':
                    response_data[epg['epg_name']] = epg_resp
                    continue

                tag = GenericClass('tagInst')
                tag.attributes.dn = f'{epg_resp["fvAEPg"]["attributes"]["dn"]}/tag-{epg["tag"]}'
                tag.attributes.status = 'created'
                r = apic.post(configuration=tag.json())

                response_data[epg['epg_name']] = r.reason

        return 200, response_data

    @classmethod
    def assign_epg_to_aep(cls, env: str, mapping: dict):
        with cls(env=env) as apic:
            aep = mapping['AEP']
            tenant = mapping['Tenant']
            ap = mapping['AP']
            epg = mapping['EPG']

            epg_data = EPG.load(apic.collect_epgs(tenant, epg))
            if ap not in epg_data.attributes.dn:
                return 500, ['EPG lookup error']

            # epg_vlan_data = apic.get_vlan_data(epg=epg)
            epg_vlan_data = apic.get_vlan_data(dn=f'uni/tn-{tenant}/ap-{ap}/epg-{epg}')

            if epg_vlan_data == list():
                vlan = apic.get_next_vlan()
            elif len(epg_vlan_data) > 1:
                try:
                    vlan = next(key for vlan in epg_vlan_data for key in vlan
                                if 'aep-Placeholders' in vlan[key]['AEPs'])
                except StopIteration:
                    return 500, [
                        'API could not determine VLAN to use as multiple are in use and designation is not defined '
                        'on aep-Placholders']
            else:
                vlan = next(key for key in epg_vlan_data[0])

            aep_c = AEP()
            aep_c.attributes.name = aep
            aep_c.attributes.status = 'modified'
            aep_c.add_epg(epg_dn=epg_data.attributes.dn, encap=vlan)

            r = apic.post(configuration=aep_c.json(), uri=aep_c.post_uri)

            if not r.ok:
                return r.status_code, r.json()

        return 200, {'EPG': epg, 'VLAN': vlan}

    def get_aep_usage(self, aep_name: str):
        # Collect Physical Interfaces using the defined AEP; Class l1RsAttEntityPCons
        interfaces = [GenericClass.load(intf) for intf in self.get(f'/api/class/l1RsAttEntityPCons.json?'
                                                                   f'query-target-filter=wcard(l1RsAttEntityPCons.tDn,'
                                                                   f'"{aep_name}")').json()['imdata']]

        # Extract switch information and Construct response data for interfaces
        response_data = {}

        for interface in interfaces:
            switch = int(re.search(r'node-(\d+)', interface.attributes.dn)[1])
            port = int(re.search(r'\d+$', interface.attributes.parentSKey).group())

            if switch not in response_data.keys():
                response_data[switch] = [port]
            else:
                response_data[switch].append(port)

        for key in response_data:
            response_data[key].sort()

        return response_data

    def update_static_route(self, tenant: str, cidr: str, new_next_hop: str):
        routes = self.get(f'/api/class/ipRouteP.json'
                          f'?query-target-filter=wcard(ipRouteP.dn,"{tenant}")&rsp-subtree=full'
                          f'&rsp-prop-include=config-only').json()['imdata']
        routes = [GenericClass.load(_) for _ in routes]
        routes = [_ for _ in routes if _.attributes.ip == cidr]

        result = []

        for route in routes:
            # Copy existing next hop
            nnh = GenericClass.load(route.children[0].json())

            # Set original next hop to be deleted
            route.children[0].attributes.status = 'deleted'

            # Create updated next hop
            nnh.attributes.nhAddr = new_next_hop
            nnh.attributes.status = 'created'
            route.children.append(nnh)

            r = self.post(route.json())

            result.append({'status': r.status_code, 'post_data': route.json()})

        return 200, result

    @classmethod
    def verify_leaf_uplinks(cls, env: str):
        with cls(env=env) as apic:
            not_compliant = []

            nodes = apic.collect_nodes()
            nodes = [FabricNode.load(x) for x in nodes]
            leafs = [x for x in nodes if x.attributes.role == 'leaf']
            spines = [x for x in nodes if x.attributes.role == 'spine']
            spines = [x.attributes.name for x in spines]

            lldp_neighbors = apic.get('/api/class/lldpAdjEp.json').json()['imdata']
            lldp_neighbors = [GenericClass.load(x) for x in lldp_neighbors]

            for leaf in leafs:
                neighbors = [x for x in lldp_neighbors if f'node-{leaf.attributes.id}' in x.attributes.dn]
                neighbors = [x for x in neighbors for spine in spines
                             if spine in x.attributes.sysName]
                neighbors = [f'{x.attributes.sysName} - {x.attributes.portIdV}' for x in neighbors]
                neighbors.sort()

                if len(neighbors) % 2:
                    not_compliant.append({'Leaf': leaf.attributes.name, 'Established_Uplinks': list(neighbors)})

            if not_compliant:
                return not_compliant
            else:
                return True

    def mark_unused_port_profiles(self):
        used_profiles = self.get('/api/class/infraRsAccPortP.json').json()['imdata']
        used_profiles = [re.search(r'accportprof-(.*)$', APICObject.load(_).attributes.tDn).group(1)
                         for _ in used_profiles]

        all_profiles = self.get('/api/class/infraAccPortP.json?%s' % CONFIG_ONLY).json()['imdata']
        all_profiles = [APICObject.load(_).attributes.name for _ in all_profiles]

        unused_profiles = set(all_profiles).difference(set(used_profiles))

        for p in unused_profiles:
            ifp = APICObject.load(self.collect_interface_profiles(name=p))
            ifp.attributes.descr = (ifp.attributes.descr if ifp.attributes.descr.startswith('UNUSED') else
                                    f'UNUSED: {ifp.attributes.descr}')
            ifp.attributes.status = 'modified'

            self.post(configuration=ifp.self_json(), uri=ifp.post_uri)

        return None

    def remove_maintenance_policies(self):
        # Just delete all maintenance groups and policies
        groups = [APICObject.load(_) for _ in self.collect(maintMaintGrp='')]
        policies = [APICObject.load(_) for _ in self.collect(maintMaintP='')]

        for _ in groups + policies:
            _.delete()
            self.post(_.json())

    def verify_maintenance_policies_and_groups(self):
        def check_group(name: str):
            if not self.dn_exists(maintMaintGrp='uni/fabric/maintgrp-%s' % name):
                group = MaintenanceGroup()
                group.attributes.dn = 'uni/fabric/maintgrp-%s' % name
                group.use_maintenance_policy(name=name)
                group.create()

                self.post(group.json())

        def check_pol(name: str):
            if not self.dn_exists(maintMaintP='uni/fabric/maintpol-%s' % name):
                pol = MaintenancePolicy()
                pol.name = name
                pol.set_firmware_version(version=(self.version if name == STAGING else self.leaf_version))
                pol.attributes.__setattr__('ignoreCompat', ('yes' if name == STAGING else 'no'))

                self.post(pol.json())

        # Check to see if Spine-Odds Policy exists
        check_pol(SPINE_ODDS)
        check_group(SPINE_ODDS)

        # Check to see if Spine-Evens Policy exists
        check_pol(SPINE_EVENS)
        check_group(SPINE_EVENS)

        # Check to see if Odds Policy exists
        check_pol(ODDS)
        check_group(ODDS)

        # Check to see if Evens Policy exists
        check_pol(EVENS)
        check_group(EVENS)

        # Check to see if OOB Policy exists
        check_pol(OOB)
        check_group(OOB)

        # Check to see if Staging Policy exists
        check_pol(STAGING)
        check_group(STAGING)

        # Enforce that 8 hour scheduler is assigned to Staging maintenance policy
        schedule = {
            'maintRsPolScheduler': {
                'attributes': {
                    'dn': 'uni/fabric/maintpol-%s/rspolScheduler' % STAGING,
                    'tnTrigSchedPName': 'EveryEightHours',
                    'status': 'created,modified'
                }
            }
        }
        self.post(schedule)

        return None

    def assign_nodes_to_maintenance_groups(self):
        def gen_blk(group: str, node_id: str):
            # Generate and return a fabricNodeBlk object using the group and node ID provided
            b = FabricNodeBlock()
            b.group = group
            b.node = node_id
            b.attributes.to_ = node_id
            b.attributes.from_ = node_id
            b.create()

            return b

        self.verify_maintenance_policies_and_groups()

        # # Delete all existing node blocks in maintenance groups
        # blocks = [APICObject.load(_) for _ in self.collect(fabricNodeBlk='')]
        # for block in blocks:
        #     block.delete()
        #     self.post(block.json())
        #
        # Collect all nodes
        nodes = [APICObject.load(_) for _ in self.collect_nodes()]

        data_leaf_nodes = [_ for _ in nodes if 100 < int(_.attributes.id) < 200] + \
                          [_ for _ in nodes if 500 < int(_.attributes.id) < 600] + \
                          [_ for _ in nodes if 2100 < int(_.attributes.id) < 2200]  # Because of multi-pod in QOL
        oob_leaf_nodes = [_ for _ in nodes if 300 < int(_.attributes.id) < 400] + \
                         [_ for _ in nodes if 600 < int(_.attributes.id) < 700] + \
                         [_ for _ in nodes if 2300 < int(_.attributes.id) < 2400]  # Because of multi-pod in QOL
        spine_nodes = [_ for _ in nodes if 200 < int(_.attributes.id) < 300] + \
                      [_ for _ in nodes if 2200 < int(_.attributes.id) < 2300]  # Because of multi-pod in QOL

        for node in data_leaf_nodes:
            if int(node.attributes.id) % 2:
                # Assign to the Odds maintenance group
                blk = gen_blk(ODDS, node.attributes.id)
                self.post(blk.json())
            else:
                # Assign to the Evens maintenance group
                blk = gen_blk(EVENS, node.attributes.id)
                self.post(blk.json())

        for node in oob_leaf_nodes:
            blk = gen_blk(OOB, node.attributes.id)
            self.post(blk.json())

        for node in spine_nodes:
            if int(node.attributes.id) % 2:
                # Assign to the Spine-Odds maintenance group
                blk = gen_blk(SPINE_ODDS, node.attributes.id)
                self.post(blk.json())
            else:
                # Assign to the Spine-Evens maintenance group
                blk = gen_blk(SPINE_EVENS, node.attributes.id)
                self.post(blk.json())

        return None

    def reclaim_interfaces(self, profile_name: str, interfaces: str):
        interfaces = intf_range(interfaces.split(','))
        proceed = True
        active_interfaces = []

        # Check to make sure ports are operationally down
        # Determine nodes to check interfaces on
        n_profiles = self.get('/api/mo/uni/infra/accportprof-%s.json?query-target=subtree&'
                              'target-subtree-class=infraRtAccPortP' % profile_name).json()['imdata']
        n_profiles = [APICObject.load(_).attributes.tDn for _ in n_profiles]
        n_profiles = [re.search(r'nprof-(.*)', _).group(1) for _ in n_profiles]

        node_blocks = []
        for name in n_profiles:
            node_blocks += self.get('/api/mo/uni/infra/nprof-%s.json?query-target=subtree&'
                                    'target-subtree-class=infraNodeBlk' % name).json()['imdata']

        node_blocks = [APICObject.load(_) for _ in node_blocks]

        nodes = set()
        for node in node_blocks:
            nodes.add(node.attributes.from_)
            nodes.add(node.attributes.to_)

        if len(nodes) > 2:
            return 400, {'message': 'Reclamation process aborted.  The specified policy is used by more than 2 '
                                    'switches',
                         'nodes': list(nodes)}

        for node in nodes:
            for interface in interfaces:
                # Interface status obtained from ethpm.PhysIf.operSt
                status = APICObject.load(self.get('/api/mo/topology/pod-1/node-%s/sys/phys-[eth1/%d]/phys'
                                                  '.json' % (node, interface)).json()['imdata'][0])

                if status.attributes.operSt == 'up':
                    # Interface is active, abort reclamation
                    proceed = False
                    active_interfaces.append([node, interface])

        if not proceed:
            return 400, {'message': 'Reclamation process aborted.  One or more interfaces were found active',
                         'active_interfaces': active_interfaces}

        # Collect port blocks defined within port profile
        blocks = self.get('/api/mo/uni/infra/accportprof-%s.json'
                          '?query-target=subtree&target-subtree-class=infraPortBlk'
                          '&rsp-prop-include=config-only' % profile_name).json()['imdata']
        blocks = [InterfaceBlock.load(_) for _ in blocks]

        # Extract interface selectors from the block distinguished names
        sels = set()

        for block in blocks:
            sels.add(re.search(r'hports-([^/]*)-typ-', block.attributes.dn).group(1))

        # This process will create a dictionary of interface selectors and the ports defined in each of them
        sels_ranges = dict()

        for block in blocks:
            sel = re.search(r'hports-([^/]*)-typ-', block.attributes.dn).group(1)
            if sels_ranges.get(sel):
                sels_ranges[sel] += intf_range(['%s-%s' % (block.attributes.fromPort, block.attributes.toPort)])
            else:
                sels_ranges[sel] = intf_range(['%s-%s' % (block.attributes.fromPort, block.attributes.toPort)])

        for k in sels_ranges:
            sels_ranges[k].sort()

        print('Before:', sels_ranges)

        # Break up port blocks given the interfaces provided
        for k, v in sels_ranges.items():
            v_set = set(v)
            i_set = set(interfaces)
            diff = v_set.difference(i_set)

            sels_ranges[k] = [list(map(itemgetter(1), g)) for k, g in groupby(enumerate(list(diff)),
                                                                              lambda x: x[0] - x[1])]

        print('After', sels_ranges)

        prof = InterfaceProfile()
        prof.name = profile_name

        for sel in sels_ranges:
            sele = InterfaceSelector()
            sele.attributes.name = sel

            prof.children.append(sele)

            for range_ in sels_ranges[sel]:
                new_block = InterfaceBlock()
                new_block.name = ''.join(random.choices(string.ascii_lowercase + string.digits, k=16))
                new_block.attributes.name = new_block.name
                new_block.attributes.fromPort = str(range_[0])
                new_block.attributes.toPort = str(range_[-1])
                new_block.attributes.descr = sele.attributes.name
                new_block.create()

                sele.children.append(new_block)

        # Stage ALL of the old blocks to be deleted
        for block in blocks:
            block.attributes.status = 'deleted'
            sel = re.search('hports-([^/]*)-typ-', block.attributes.dn).group(1)
            sele = next(c for c in prof.children if c.attributes.name == sel)
            sele.children.insert(0, block)

        r = self.post(prof.json(), uri='/api/mo/uni/infra.json')

        return r.status_code, {'message': ('Interfaces have been defaulted' if r.ok else 'An error occurred'),
                               'response_data': r.json(), 'submitted_data': prof.json()}

    def assign_vlan_to_aep(self, vlan: int or str, aep: str):
        vlan = int(vlan)
        vlan_data = self.get_vlan_data(vlan=vlan)

        if not isinstance(vlan_data, dict):
            return 400, {'message': 'VLAN ID has no associated EPG'}

        if aep not in vlan_data[vlan]['AEPs']:
            if len(vlan_data[vlan]['Consumers']) > 1:
                return 400, {'message': 'There is a VLAN conflict for this VLAN ID.  Contact a network engineer to '
                                        'resolve'}
            tenant, ap, epg = re.search(r'tn-([^/]+)/ap-([^/]+)/epg-([^/]+)', vlan_data[vlan]['Consumers'][0]).groups()
            func = InfraRsFuncToEpg()
            func.aep = aep
            func.tenant = tenant
            func.app_profile = ap
            func.epg = epg
            func.attributes.encap = 'vlan-%d' % vlan

            response = self.post(func.json())

            return response.status_code, {'EPG': vlan_data[vlan]['Consumers'][0], 'APIC Response': response.json()}

        return 400, {'message': 'VLAN %d is already associated to AEP %s' % (vlan, aep)}

    def remove_unused_epgs_bds_subnets(self):
        # Collect all EPGs from production and ADMZ tenants  - Omitting HCADR for now
        epgs = [APICObject.load(epg) for epg in self.collect_epgs()]
        epgs = [epg for epg in epgs if re.search(r'tn-([^/]+)', epg.attributes.dn).group(1) in [self.env.Tenant,
                                                                                                self.env.ADMZTenant]]
        all_epgs = set(epg.attributes.dn for epg in epgs)

        # Collect all endpoints, fvCEp
        endpoints = [APICObject.load(_) for _ in self.collect_eps()]
        # Filter to endpoints known via EPGs
        endpoints = [_ for _ in endpoints if re.match(r'uni/tn-[^/]+/ap-[^/]+/epg-[^/]+', _.attributes.dn)]

        ipaddresses = set(_.attributes.ip for _ in endpoints if 'ip' in _.attributes.json().keys()
                          if bool(int(IPv4Address(_.attributes.ip))))

        # Collect all subnets, fvSubnet
        subnets = [APICObject.load(_) for _ in self.collect_subnets()]
        subnets = [_ for _ in subnets if re.search(r'tn-([^/]+)', _.attributes.dn).group(1) in [self.env.Tenant,
                                                                                                self.env.ADMZTenant]]
        # Filter out L3Out based networks
        subnets = [_ for _ in subnets if _.class_ == Subnet.class_]

        all_subnets = set(_.attributes.dn for _ in subnets)

        networks = [(IPv4Network(_.attributes.ip, strict=False), _.attributes.dn) for _ in subnets]

        used_subnets = set()
        for ip in ipaddresses:
            for network, dn in networks:
                if network.overlaps(IPv4Network(ip)):
                    used_subnets.add(dn)

        used_epgs = set(re.search(r'uni/tn-[^/]+/ap-[^/]+/epg-[^/]+', _.attributes.dn).group() for _ in endpoints)

        unused_epgs = all_epgs.difference(used_epgs)

        # Collect all bridge domains from production and ADMZ tenants
        bds = [APICObject.load(_) for _ in self.get('/api/class/fvBD.json').json()['imdata']]
        bds = [bd for bd in bds if re.search(r'tn-([^/]+)', bd.attributes.dn).group(1) in [self.env.Tenant,
                                                                                           self.env.ADMZTenant]]
        all_bds = set(bd.attributes.dn for bd in bds)

        used_bds = [APICObject.load(_).attributes.tDn for _ in self.get('/api/class/fvRsBd.json').json()['imdata']]

        unused_bds = all_bds.difference(set(used_bds))

        unused_subnets = all_subnets.difference(used_subnets)

        # return unused_epgs, unused_bds, (unused_subnets, used_subnets)
        return 200, {'unused_epgs': list(unused_epgs), 'unused_bds': list(unused_bds),
                     'unused_subnets': list(unused_subnets)}

    # def create_new_epg_tf(self, app_profile: str, app_instance_name: str, description: str, num_ips_needed: int) -> \
    #         Tuple[int, str]:
    #
    #     with BIG() as big:
    #         subnet = big.assign_next_network_from_list(block_list=self.env.Subnets, no_of_ips=num_ips_needed,
    #                                                    name=app_instance_name, coid=int(self.env.COID),
    #                                                    asn=int(self.env.ASN))
    #         network = IPv4Network(subnet.properties['CIDR'])
    #         gateway = network.network_address + 1
    #
    #     aps = [APICObject.load(a) for a in self.collect_aps()]
    #
    #     try:
    #         ap = next(a for a in aps if a.attributes.annotation.__dict__.get('tfref') == app_profile.lower())
    #     except StopIteration:
    #         return 404, 'Terraform does not yet support this application profile.  Contact William Duhe for support'
    #
    #     ap_ref = '.'.join([AP.tf_resource, ap.attributes.annotation.tfref, 'id'])
    #
    #     # Create TF file content
    #     tf_name = re.sub(r'\W', '_', app_instance_name).lower()
    #
    #     tfbd = Resource(BD.tf_resource, tf_name,
    #                     tenant_dn='local.tenantdn',
    #                     name=f'"bd-{app_instance_name}"',
    #                     description=f'"{description}"',
    #                     relation_fv_rs_bd_to_out='[local.l3outcoredn]',
    #                     relation_fv_rs_ctx='local.vrfdn',
    #                     arp_flood='"no"    # Options: yes | no',
    #                     unicast_route='"yes"    # Options: yes | no',
    #                     unk_mac_ucast_act='"proxy"    # Options: proxy | flood')
    #
    #     tfepg = Resource(EPG.tf_resource, tf_name,
    #                      application_profile_dn=ap_ref,
    #                      name=f'"epg-{app_instance_name}"',
    #                      description=f'"{description}"',
    #                      relation_fv_rs_bd=tfbd.ref_id())
    #
    #     tfsubnet = Resource(Subnet.tf_resource, tf_name,
    #                         parent_dn=tfbd.ref_id(),
    #                         # ip='"192.168.2.1/29"',
    #                         ip=f'"{gateway}/{network.prefixlen}"',
    #                         scope=json.dumps(['public']))
    #
    #     tfepgdomain = Resource(FvRsDomAtt.tf_resource, tf_name,
    #                            application_epg_dn=tfepg.ref_id(),
    #                            tdn='local.physicaldomaindn')
    #
    #     resources = ['', tfbd, tfepg, tfsubnet, tfepgdomain]
    #
    #     # Add the EPG to the AEP with the VLAN returned
    #     epg_dn = f'uni/tn-{self.env.Tenant}/ap-{ap.attributes.name}/epg-{app_instance_name}'
    #
    #     vlan_assignment = InfraRsFuncToEpg(encap=f'vlan-{self.get_next_vlan()}', tDn=epg_dn)
    #     vlan_assignment.create()
    #
    #     # POST the encapsulation to aep-Placeholders
    #     # self.post(configuration=vlan_assignment.json(),
    #     #           uri='/api/mo/uni/infra/attentp-aep-Placeholders/gen-default.json')
    #
    #     content = '\n'.join([str(x) for x in resources])
    #
    #     return 200, content

    # def create_custom_epg_tf(self, app_profile: str, app_instance_name: str, description: str, network_cidr: str) -> \
    #         Tuple[int, str]:
    #
    #     network = IPv4Network(network_cidr, strict=False)
    #     gateway = network.network_address + 1
    #
    #     aps = [APICObject.load(a) for a in self.collect_aps()]
    #
    #     try:
    #         ap = next(a for a in aps if a.attributes.annotation.__dict__.get('tfref') == app_profile.lower())
    #     except StopIteration:
    #         return 404, 'Terraform does not yet support this application profile.  Contact William Duhe for support'
    #
    #     ap_ref = '.'.join([AP.tf_resource, ap.attributes.annotation.tfref, 'id'])
    #
    #     # Create TF file content
    #     tf_name = re.sub(r'\W', '_', app_instance_name).lower()
    #
    #     tfbd = Resource(BD.tf_resource, tf_name,
    #                     tenant_dn='local.tenantdn',
    #                     name=f'"bd-{app_instance_name}"',
    #                     description=f'"{description}"',
    #                     relation_fv_rs_bd_to_out='[local.l3outcoredn]',
    #                     relation_fv_rs_ctx='local.vrfdn',
    #                     arp_flood='"no"    # Options: yes | no',
    #                     unicast_route='"yes"    # Options: yes | no',
    #                     unk_mac_ucast_act='"proxy"    # Options: proxy | flood')
    #
    #     tfepg = Resource(EPG.tf_resource, tf_name,
    #                      application_profile_dn=ap_ref,
    #                      name=f'"epg-{app_instance_name}"',
    #                      description=f'"{description}"',
    #                      relation_fv_rs_bd=tfbd.ref_id())
    #
    #     tfsubnet = Resource(Subnet.tf_resource, tf_name,
    #                         parent_dn=tfbd.ref_id(),
    #                         ip=f'"{gateway}/{network.prefixlen}"',
    #                         scope=json.dumps(['public']))
    #
    #     tfepgdomain = Resource(FvRsDomAtt.tf_resource, tf_name,
    #                            application_epg_dn=tfepg.ref_id(),
    #                            tdn='local.physicaldomaindn')
    #
    #     resources = ['', tfbd, tfepg, tfsubnet, tfepgdomain]
    #
    #     # Add the EPG to the AEP with the VLAN returned
    #     epg_dn = f'uni/tn-{self.env.Tenant}/ap-{ap.attributes.name}/epg-{app_instance_name}'
    #
    #     vlan_assignment = InfraRsFuncToEpg(encap=f'vlan-{self.get_next_vlan()}', tDn=epg_dn)
    #     vlan_assignment.create()
    #
    #     # POST the encapsulation to aep-Placeholders
    #     # self.post(configuration=vlan_assignment.json(),
    #     #           uri='/api/mo/uni/infra/attentp-aep-Placeholders/gen-default.json')
    #
    #     content = '\n'.join([str(x) for x in resources])
    #
    #     return 200, content

    # @classmethod
    # def move_network(cls, source_az: str, dest_az: str, cidr: str, app_profile: str, inst_name: str):
    #     cidr = IPv4Network(cidr, strict=False)
    #     s_apic = cls(env=source_az)
    #     d_apic = cls(env=dest_az)
    #
    #     ssnap = s_apic.snapshot(descr=f'{inst_name}_move')
    #     dsnap = d_apic.snapshot(descr=f'{inst_name}_move')
    #
    #     if not ssnap or not dsnap:
    #         return 400, {'message': 'Snapshot creation failed in one of the environments, not proceeding'}
    #
    #     subnets = [APICObject.load(_)
    #                for _ in s_apic.get(f'/api/mo/uni/tn-{s_apic.env.Tenant}.json'
    #                                    f'?query-target=subtree&target-subtree-class=fvSubnet').json()['imdata']]
    #     subnets += [APICObject.load(_)
    #                 for _ in s_apic.get(f'/api/mo/uni/tn-{s_apic.env.ADMZTenant}.json'
    #                                     f'?query-target=subtree&target-subtree-class=fvSubnet').json()['imdata']]
    #
    #     subnets = [[IPv4Network(_.attributes.ip, strict=False), _] for _ in subnets]
    #
    #     # Get and process subnet information
    #     try:
    #         network, subnet = next([n, s] for n, s in subnets if n.overlaps(cidr) and n.prefixlen == cidr.prefixlen)
    #         subnet.remove_admin_props()
    #         s_subnet = Subnet.load(subnet.json())
    #         s_subnet.delete()
    #         subnet.attributes.__delattr__('dn')
    #         subnet.create()
    #     except StopIteration:
    #         return 404, {'message': 'The prefix does not exist in the source environment'}
    #
    #     tenant_name, bd_name = re.search('tn-([^/]+)/BD-([^/]+)', s_subnet.attributes.dn).groups()
    #
    #     # Define Tenant
    #     tenant = Tenant(name=(d_apic.env.Tenant if s_apic.env.Tenant == tenant_name else d_apic.env.ADMZVRF))
    #     tenant.modify()
    #
    #     # Get and setup Bridge Domain
    #     bd = APICObject.load(s_apic.collect_bds(tn=tenant_name, bd=bd_name))
    #     bd.children = []
    #     bd.attributes.__delattr__('dn')
    #     bd.attributes.name = f'bd-{inst_name}'
    #     bd.use_vrf(name=(d_apic.env.VRF if s_apic.env.Tenant == tenant_name else d_apic.env.ADMZVRF))
    #     bd.create_modify()
    #     if s_apic.env.Tenant == tenant_name:
    #         bd.to_l3_out(d_apic.env.L3OutCore)
    #
    #     # Generate EPG
    #     epg = EPG(name=f'epg-{inst_name}')
    #     epg.domain(d_apic.env.PhysicalDomain)
    #     epg.assign_bd(name=bd.attributes.name)
    #     epg.create_modify()
    #
    #     # Generate AP
    #     ap = AP()
    #     ap.tenant = (d_apic.env.Tenant if s_apic.env.Tenant == tenant_name else d_apic.env.ADMZTenant)
    #     ap.attributes.name = f'ap-{app_profile}'
    #     ap.create_modify()
    #
    #     # Setup structure
    #     tenant.children = [bd, ap]
    #     bd.children.append(subnet)
    #     ap.children = [epg]
    #
    #     # Create the objects in the destination environment
    #     print(json.dumps(tenant.json()))
    #     # r = d_apic.post(tenant.json())
    #     # if not r.ok:
    #     #     return 400, {'message': 'Creating the network in the target environment failed. Revert Snapshots'}
    #
    #     # Delete the Subnet
    #     s_subnet.delete()
    #     print(json.dumps(subnet.json()))
    #     # r = s_apic.post(s_subnet.json())
    #     # if not r.ok:
    #     #     return 400, {'message': 'Deleting the network from the source environment failed. Revert Snapshots'}
    #
    #     return None

    def document_app_instances(self):
        # Collect subnets
        subnets = self.get(f'/api/mo/uni/tn-{self.env.Tenant}.json'
                           f'?query-target=subtree&target-subtree-class={Subnet.class_}').json()['imdata']
        subnets = [APICObject.load(_) for _ in subnets]

        # Get all EPs and filter out EPs without IP addresses and non-EPG EPs
        all_eps = [APICObject.load(_) for _ in self.get('/api/class/fvCEp.json').json()['imdata']]
        all_eps = [_ for _ in all_eps if _.attributes.ip and '/epg-' in _.attributes.dn]

        gh = GithubAPI()

        count = 0

        for subnet in subnets:
            count += 1
            print(f'Processing subnet {count} of {len(subnets)}', end='\r')
            cidr = IPv4Network(subnet.attributes.ip, strict=False)
            eps = [_ for _ in all_eps
                   if cidr.overlaps(IPv4Network(_.attributes.ip))]  # Filter EPs to match the subnet of interest

            if eps:
                orig_dn = EPG_DN_SEARCH.search(eps[0].attributes.dn).group()
                epg_tenant, epg_ap, epg_name = EPG_DN_SEARCH.search(eps[0].attributes.dn).groups()
                inst_name = AppInstance.format_name(epg_name)
                # inst_name = re.sub(r'^epg-', '', epg_name).lower().replace('-', '_')
                # inst_name = f'{self.env.DataCenter}_{inst_name}'.lower()
                application = AppInstance.format_name(epg_ap)

                # Collect bridge domain settings
                rsbd = APICObject.load(self.get(f'/api/mo/{orig_dn}/rsbd.json').json()['imdata'])
                bd = APICObject.load(self.collect_bds(tn=self.env.Tenant, bd=rsbd.attributes.tnFvBDName))

                subnet.remove_admin_props()
                # Do not document distinguished name of the subnet
                for attr in ['dn']:
                    subnet.attributes.__delattr__(attr)

                inst = AppInstance(epgName=epg_name, application=application, currentAZ=self.__str__(),
                                   bdName=bd.attributes.name, apName=epg_ap, name=inst_name,
                                   tenant=next(_ for _ in dir(self.env)
                                               if self.env.__getattribute__(_) == epg_tenant),
                                   networks={subnet.attributes.ip: subnet.attributes.json()},
                                   bdSettings=bd.attributes.json())

                if gh.file_exists(inst.path()):
                    app_inst = AppInstance.load(inst.path())
                    if app_inst.content() != inst.content():
                        app_inst.update(**inst.json())
                        app_inst.store()
                    else:
                        continue
                else:
                    gh.add_file(file_path=inst.path(), message=f'{inst}_add', content=inst.content())

        print()
        return None

    def find_lldp_neighbors(self, mac_addresses: str) -> Tuple[int, list]:
        mac_addresses = format_mac_addresses(mac_addresses.lower().split(','))

        lldp_neigh = [APICObject.load(_) for _ in self.get('/api/class/lldpAdjEp.json').json()['imdata']]
        lldp_neigh = [neigh for neigh in lldp_neigh if neigh.attributes.portIdV in mac_addresses]

        lldp_neigh = [{
            'node': re.search(r'node-(\d+)', neigh.attributes.dn).group(1),
            'port': re.search(r'eth\d+/\d+', neigh.attributes.dn).group(),
            'sysName': (neigh.attributes.sysName if neigh.attributes.sysName else 'unspecified'),
            'chassisID': neigh.attributes.chassisIdV,
            'portIdV': neigh.attributes.portIdV,
            'mgmtIp': neigh.attributes.mgmtIp
        } for neigh in lldp_neigh]

        return 200, lldp_neigh


class AppInstance:
    ADMZTENANT = 'ADMZTenant'
    TENANT = 'Tenant'
    apName: str
    bdName: str
    epgName: str
    application: str
    tenant: str
    currentAZ: APIC
    originAZ: APIC
    bdSettings: dict
    name: str
    networks: dict
    createTime: datetime
    modifiedTime: datetime
    __naming_attrs = ['apName', 'bdName', 'epgName']
    __time_attrs = ['createTime', 'modifiedTime']
    __apic_attrs = ['currentAZ', 'originAZ']
    __str_attrs = ['apName', 'bdName', 'epgName', 'epgDn', 'tenant', 'application', 'name']
    __dict_attrs = ['networks', 'bdSettings']
    __int_attrs = []
    GITHUB_PATH = 'applications'
    ISSUES_PATH = 'pyapis/appinst/issues'

    def __init__(self, **kwargs):
        assert 'name' and 'application' and 'tenant' in kwargs

        self.__all_attrs = self.__time_attrs + self.__apic_attrs + self.__dict_attrs + self.__str_attrs + \
            self.__int_attrs + self.__naming_attrs

        for attr in self.__dict_attrs + self.__str_attrs + self.__naming_attrs:
            if attr == 'name' or attr == 'application':
                value = self.format_name(kwargs.get(attr))
                self.__setattr__(attr, value)
                continue
            self.__setattr__(attr, kwargs.get(attr))

        for attr in self.__time_attrs:
            v = kwargs.get(attr)
            if v:
                self.__setattr__(attr, datetime.fromisoformat(v))
            else:
                self.__setattr__(attr, None)

        for attr in self.__apic_attrs:
            v = kwargs.get(attr)
            if v:
                self.__setattr__(attr, APIC(env=v))
            else:
                self.__setattr__(attr, None)

        if not self.createTime:
            now = datetime.now()
            self.createTime = now
            self.modifiedTime = now

        if not self.originAZ:
            self.originAZ = self.currentAZ

        if not self.bdSettings:
            self.bdSettings = {}

        if not self.networks:
            self.networks = {}

        self.__activate()

        if not self.name.startswith(f'{self.format_name(self.originAZ.env.Name)}_'.lower()):
            self.name = f'{self.format_name(self.originAZ.env.Name)}_{self.name}'.lower()

    def __str__(self):
        self.__activate()
        return self.name.lower()

    def __activate(self) -> None:
        for attr in self.__apic_attrs:
            if self.__getattribute__(attr) and not isinstance(self.__getattribute__(attr), APIC):
                self.__setattr__(attr, APIC(env=self.__getattribute__(attr)))

    def path(self):
        self.__activate()
        return f'{self.GITHUB_PATH}/{self.application}/{self}'

    @classmethod
    def load(cls, app_inst_path: str):
        if not app_inst_path.startswith(f'{cls.GITHUB_PATH}/'):
            app_inst_path = f'{cls.GITHUB_PATH}/{app_inst_path}'
        gh = GithubAPI()
        if gh.file_exists(app_inst_path):
            c = gh.get_file_content(app_inst_path)
            inst = cls(**json.loads(c))
            return inst
        else:
            raise NameError(f'The application instance does not exist: {app_inst_path}')

    def update(self, **kwargs):
        pot_inst = AppInstance(**kwargs)
        if pot_inst.epg_dn(override=True) != self.epg_dn(override=True):
            raise NameError('The EPG distinguished name in the update data is not consistent with the AppInstance EPG '
                            'distinguished name: %s != %s' % (pot_inst.epg_dn(), self.epg_dn()))

        for k, v in kwargs.items():
            if k not in self.__all_attrs:
                continue
            elif k in self.__dict_attrs:
                self.__dict__.get(k).update(v)
            elif k == 'originAZ':
                continue  # Do not update originAZ using this method
            else:
                self.__setattr__(k, v)

        self.modifiedTime = datetime.now()

    def store(self):
        self.modifiedTime = datetime.now()

        gh = GithubAPI()

        if gh.file_exists(self.path()):
            gh.update_file(file_path=self.path(), message=f'{self}_update', content=self.content())
        else:
            gh.add_file(file_path=self.path(), message=f'{self}_add', content=self.content())

    def refactor(self, new_application: str=None, new_name: str=None):
        """Renames and/or Moves an instance within GitHub based on the provided criteria"""
        gh = GithubAPI()

        # Get original values
        original_path = self.path()
        original_dn = self.epg_dn()
        orig_tn, orig_ap, orig_epg = EPG_DN_SEARCH.search(original_dn).groups()

        if new_application:
            self.application = self.format_name(new_application)

        if new_name:
            if not new_name.startswith(f'{self.format_name(self.originAZ.env.Name)}_'.lower()):
                self.name = f'{self.format_name(self.originAZ.env.Name)}_{new_name}'.lower()
            else:
                self.name = new_name

        # Check to see if the new generated dn will be different from the current generated dn. Set attributes to keep
        # original dn for APIC tasks, if so.
        if self.epg_dn() != original_dn:
            self.epgName = orig_epg
            self.apName = orig_ap

        # Delete current file and store in new location
        if original_path != self.path():
            gh.delete_file(original_path, message=f'{self.name}_refactor_to_{self.path()}')
            self.store()

    def remove(self):
        gh = GithubAPI()

        gh.delete_file(self.path(), message=f'{self}_removal')

    def json(self):
        output = {}

        # Filter out any unwanted Bridge Domain specifics
        for k in list(self.bdSettings.keys()):
            if k in ['name', 'mac', 'vmac', 'dn', 'llAddr', 'status']:
                _ = self.bdSettings.pop(k)

        for attr in self.__all_attrs:
            if attr in self.__time_attrs + self.__apic_attrs + self.__str_attrs:
                output[attr] = (str(self.__dict__.get(attr)) if self.__dict__.get(attr) else None)
            elif attr in self.__dict_attrs + self.__int_attrs:
                output[attr] = self.__dict__.get(attr)

        return output

    def content(self):
        """Returns standard output for dumping into a file"""
        return json.dumps(self.json(), indent=2, sort_keys=True)

    @staticmethod
    def format_name(name: str) -> str:
        name = re.sub(r'^(ap-|epg-|bd-)', '', name).lower()
        name = re.sub(r'\W+', '_', name)
        return name

    def epg_dn(self, override: bool=False, drt: bool=False) -> str:
        self.__activate()

        if drt:
            return f'uni/tn-{DRT_TENANT}/ap-{self.ap_name()}/epg-{self.epg_name()}'
        elif not override:
            return f'uni/tn-{self.currentAZ.env.__getattribute__(self.tenant)}/ap-{self.ap_name()}/' \
                   f'epg-{self.epg_name()}'
        else:
            return f'uni/tn-{self.currentAZ.env.__getattribute__(self.tenant)}/ap-{self.application}/epg-{self}'

    def bd_dn(self, override: bool=False, drt: bool=False) -> str:
        self.__activate()

        if drt:
            return f'uni/tn-{DRT_TENANT}/BD-{self.bd_name()}'
        elif override:
            return f'uni/tn-{self.currentAZ.env.__getattribute__(self.tenant)}/BD-{self}'
        else:
            return f'uni/tn-{self.currentAZ.env.__getattribute__(self.tenant)}/BD-{self.bd_name()}'

    def subnet_dn(self, network: str, drt: bool=False):
        return f'{self.bd_dn(drt=drt)}/subnet-[{network}]'

    def bd_name(self):
        return self.bdName if self.bdName else str(self)

    def epg_name(self):
        return self.epgName if self.epgName else str(self)

    def ap_name(self):
        return self.apName if self.apName else self.application

    def base_name(self):
        return re.sub(f'({self.format_name(self.originAZ.env.Name)}_|{self.originAZ.env.DataCenter}_)+', '', self.name,
                      flags=re.IGNORECASE)

    def placeholder_mapping(self, drt: bool=False) -> dict:
        self.__activate()
        t, a, e = EPG_DN_SEARCH.search(self.epg_dn()).groups()
        mapping = dict(AEP=APIC.PLACEHOLDERS, Tenant=(DRT_TENANT if drt else t), AP=a, EPG=e)
        return mapping

    def generate_config(self, origin_az: bool=False, defaults: bool=False, delete: bool=False, drt: bool=False,
                        custom_az: str=False):
        az = (APIC(env=custom_az) if custom_az else self.originAZ if origin_az else self.currentAZ)

        tenant = Tenant(name=(az.env.__getattribute__(self.tenant) if not drt else DRT_TENANT))
        tenant.create_modify()

        bd = (BD(**self.bdSettings) if self.bdSettings else BD())
        if BD().json() == bd.json():
            # Configure BD settings if none presently exist
            bd.layer3()
        bd.attributes.name = (str(self) if defaults else self.bd_name())
        bd.use_vrf((DRT_VRF if drt else az.env.VRF if self.tenant == self.TENANT else az.env.ADMZVRF))
        if self.tenant == self.TENANT and not drt:
            bd.to_l3_out(name=az.env.L3OutCore)
        bd.create_modify()

        ap = AP(name=(self.application if defaults else self.ap_name()))
        ap.create_modify()

        epg = EPG(name=(str(self) if defaults else self.epg_name()))
        epg.domain(az.env.PhysicalDomain)
        epg.assign_bd(name=bd.attributes.name)
        if drt:
            contract_cons = GenericClass('fvRsCons', tnVzBrCPName='c-drtest-advertise')
            epg.children += [contract_cons]
        if delete:
            epg.delete()
        else:
            epg.create_modify()

        for prefix, settings in self.networks.items():
            subnet = Subnet()
            subnet.attributes = (Attributes(**settings) if settings else Attributes(**subnet.attributes.json()))
            subnet.attributes.ip = prefix
            if drt:
                subnet.attributes.__setattr__('scope', 'public,shared')
            if delete:
                subnet.delete()
            else:
                subnet.create_modify()

            bd.children += [subnet]

        tenant.children = [bd, ap]
        ap.children = [epg]

        return tenant

    @classmethod
    def create_new_instance(cls, az: str, application: str, inst_name: str, no_of_ips: int, dmz: bool=False) ->\
            Tuple[int, dict]:
        """Create a new application instance"""
        # Cleanse application and inst_name input
        application = cls.format_name(application)
        inst_name = cls.format_name(inst_name)

        apic = APIC(env=az)

        app_inst = cls(application=application,
                       name=inst_name,
                       currentAZ=apic.env.Name,
                       tenant=(cls.ADMZTENANT if dmz else cls.TENANT)
                       )

        # Ensure instance does not yet exist
        gh = GithubAPI()
        if gh.file_exists(file_path=app_inst.path()):
            return 400, {'message': 'The specified application instance already exists'}

        with BIG() as big:
            network = big.assign_next_network_from_list(block_list=apic.env.Subnets, no_of_ips=no_of_ips,
                                                        name=inst_name, coid=int(apic.env.COID), asn=int(apic.env.ASN))

        if not network:
            return 400, {'message': f'{apic.env.Name} has no available network for the required number of IPs.'}

        ipnetwork = IPv4Network(network.properties['CIDR'])
        gateway = ipnetwork.network_address + 1

        n = {f'{gateway}/{ipnetwork.prefixlen}': {}}
        app_inst.networks.update(n)

        tenant = app_inst.generate_config(origin_az=True)

        bd = tenant.get_child_class(BD.class_)
        subnets = bd.get_child_class_iter(Subnet.class_)

        app_inst.bdSettings = bd.attributes.json()
        for subnet in subnets:
            app_inst.networks[subnet.attributes.ip] = subnet.attributes.json()

        # Store instance before implementing in the fabric
        app_inst.store()

        # Create the APIC objects
        r = apic.post(tenant.json())

        if not r.ok:
            return r.status_code, {'message': 'EPG creation failed, but instance creation succeeded',
                                   'instance_path': app_inst.path(),
                                   'epg_dn': app_inst.epg_dn(),
                                   'apic_status': r.status_code,
                                   'apic_json': r.json(),
                                   'configuration': tenant.json()}

        _, vlan = apic.assign_epg_to_aep(env=apic.env.Name, mapping=app_inst.placeholder_mapping())

        return 200, {'message': 'Application Instance creation successful',
                     'epg_dn': app_inst.epg_dn(),
                     'application': app_inst.application,
                     'instance_path': app_inst.path(),
                     'vlan': vlan['VLAN']}

    @classmethod
    def expand_instance(cls, application: str, inst_name: str, no_of_ips: int) -> Tuple[int, dict]:
        application = cls.format_name(application)
        inst_name = cls.format_name(inst_name)

        try:
            inst = cls.load(f'{application}/{inst_name}')
        except NameError:
            return 404, {'message': f'Application instance {application}/{inst_name} does not exist'}

        with BIG() as big:
            network = big.assign_next_network_from_list(block_list=inst.originAZ.env.Subnets, no_of_ips=no_of_ips,
                                                        name=inst_name, coid=int(inst.originAZ.env.COID),
                                                        asn=int(inst.originAZ.env.ASN))

        ipnetwork = IPv4Network(network.properties['CIDR'])
        gateway = f'{ipnetwork.network_address + 1}/{ipnetwork.prefixlen}'

        subnet = Subnet(subnet=gateway)
        inst.networks[gateway] = subnet.attributes.json()
        inst.store()

        subnet.tenant = inst.currentAZ.env.__getattribute__(inst.tenant)
        subnet.bd = inst.bd_name()
        subnet.ip_network = gateway

        # Add the new network to the current AZ
        snapshot = inst.currentAZ.snapshot(descr=f'{inst.path()}_expansion')
        response = inst.currentAZ.post(subnet.json())

        return response.status_code, {'apic_response': response.json(), 'network': str(ipnetwork),
                                      'application': inst.application, 'instance': str(inst),
                                      'availability_zone': str(inst.currentAZ), 'snapshot': snapshot}

    @classmethod
    def deploy_instance(cls, inst_path: str) -> Tuple[int, dict]:
        """Deploy instance to originAZ assuming the instance does not exist elsewhere"""
        inst = cls.load(app_inst_path=inst_path)

        tenant = inst.generate_config(origin_az=True, defaults=True)

        r = inst.originAZ.post(tenant.json())

        response = {'instance': inst.json(),
                    'apic_response': r.json(),
                    'configuration': tenant.json()
                    }

        if r.ok:
            # Update instance settings
            for attr in cls.__naming_attrs:
                inst.__setattr__(attr, None)

            inst.currentAZ = inst.originAZ

            inst.update(**inst.json())
            inst.store()

            # Assign VLAN for the EPG
            _ = APIC.assign_epg_to_aep(env=inst.originAZ.env.Name, mapping=inst.placeholder_mapping())
            return r.status_code, dict(message='Application instance deployed', **response)
        else:
            return r.status_code, dict(message='Application instance deployment failed', **response)

    @classmethod
    def move_instance(cls, app_inst_path: str, az: str) -> Tuple[int, dict]:
        app_inst_path = re.sub(r'[^/\w]', '_', app_inst_path)
        inst = cls.load(app_inst_path=app_inst_path)

        src_az = APIC(env=inst.currentAZ.env.Name)
        dst_az = APIC(env=az)

        if inst.currentAZ.__str__().lower() == az.lower():
            return 200, {'message': 'The application instance already exists in the specified AZ',
                         'application': inst.application,
                         'instance': inst.__str__(),
                         'currentAZ': inst.currentAZ,
                         'requestedAZ': az}

        # Check to see if source environment is accessible.  If so, delete source networks
        try:
            src_snapshot = inst.currentAZ.snapshot(descr=f'{inst}_move_az')
        except ConnectionError:
            src_snapshot = None

        deletions = []

        if src_snapshot:
            # Get networks from the source environment
            for network, settings in inst.networks.items():
                r = inst.currentAZ.get(f'/api/mo/{inst.subnet_dn(network)}.json')
                if int(r.json()['totalCount']):
                    subnet = APICObject.load(r.json()['imdata'])
                    subnet.delete()
                    subnet.remove_admin_props()
                    deletions.append(subnet)
                else:
                    return 404, {'message': f'Failed to collect {inst.subnet_dn(network)}'}

            # Get the EPG
            r = inst.currentAZ.get(f'/api/mo/{inst.epg_dn()}.json')
            if int(r.json()['totalCount']):
                epg = EPG()
                epg.attributes.dn = inst.epg_dn()
                epg.delete()
                deletions.append(epg)
            else:
                return 404, {'message': f'Failed to collect {inst.epg_dn()}'}

            # Determine usage of BD and whether it should be deleted
            r = inst.currentAZ.get(f'/api/class/fvRsBd.json?query-target-filter=eq(fvRsBd.tDn,"{inst.bd_dn()}")')
            if int(r.json()['totalCount']) == 1:
                rsbd = APICObject.load(r.json()['imdata'])
                if epg.attributes.dn == EPG_DN_SEARCH.search(rsbd.attributes.dn).group():
                    bd = BD(dn=inst.bd_dn())
                    bd.delete()
                    deletions.append(bd)
                else:
                    # BD usage found does not match the instance EPG; take no action
                    pass
            else:
                # BD is in use by more than one policy
                pass

        # Create the networks in the new environment
        # Change currentAZ to be the new AZ
        inst.currentAZ = dst_az
        dst_snapshot = inst.currentAZ.snapshot(descr=f'{inst}_move')
        tenant = inst.generate_config(defaults=True)  # Because of this, naming attributes should be cleared

        create_resp = dst_az.post(configuration=tenant.json())

        deletion_results = []

        if create_resp.ok:
            if src_snapshot:
                for deletion in deletions:
                    r = src_az.post(configuration=deletion.json())
                    deletion_results += [r.status_code, r.json(), deletion.json()]

        _, vlan = inst.currentAZ.assign_epg_to_aep(env=inst.currentAZ.env.Name, mapping=inst.placeholder_mapping())

        if not create_resp.ok:
            return_data = {
                'message': 'Instance move failed',
                'configuration': tenant.json(),
                'apic_json': create_resp.json(),
                'apic_status': create_resp.status_code,
                'src_snapshot': src_snapshot,
                'dst_snapshot': dst_snapshot,
                'deletions': deletion_results
            }
        else:
            if src_snapshot:
                for deletion in deletions:
                    r = src_az.post(configuration=deletion.json())

                    if deletion.class_ == Subnet.class_:
                        if not r.ok:
                            return 400, {
                                'message': 'Instance move failed; failed to delete networks from source environment',
                                'configuration': deletion.json(),
                                'apic_json': r.json(),
                                'apic_status': r.status_code,
                                'src_snapshot': src_snapshot,
                                'dst_snapshot': dst_snapshot,
                                'deletions': deletion_results
                            }

            for attr in inst.__naming_attrs:
                inst.__setattr__(attr, None)

            return_data = {
                'message': 'Move Successful',
                'configuration': tenant.json(),
                'apic_json': create_resp.json(),
                'apic_status': create_resp.status_code,
                'snapshot': src_snapshot,
                'deletions': deletion_results
            }

            # Update Github file
            inst.store()

        return create_resp.status_code, return_data

    @classmethod
    def delete_instance(cls, inst_path: str) -> Tuple[int, dict]:
        inst_path = re.sub(r'[^/\w]', '_', inst_path)
        inst = cls.load(app_inst_path=inst_path)

        # Delete EPG
        epg = APICObject.load(inst.currentAZ.get(f'/api/mo/{inst.epg_dn()}.json').json()['imdata'])
        epg.attributes.remove_admin_props()
        epg.delete()

        inst.currentAZ.post(epg.json())

        # Delete Subnets
        for network in inst.networks:
            net = IPv4Network(network, strict=False)

            subnet = APICObject.load(
                inst.currentAZ.get(f'/api/mo/fvSubnet.json?'
                                   f'query-target-filter=eq(fvSubnet.ip,"{network}")').json()['imdata']
            )
            subnet.attributes.remove_admin_props()
            subnet.delete()

            inst.currentAZ.post(subnet.json())

            # Mark network to be deleted in Proteus
            with BIG() as big:
                ip4network = big.get_network(network_cidr=str(net))
                ip4network.name = f'{ip4network.name} - DELETED FROM NETWORK'
                big.update_object(ip4network)

        # Delete Bridge Domain; if not used elsewhere
        bd = APICObject.load(inst.currentAZ.get(f'/api/mo/{inst.bd_dn()}.json').json()['imdata'])

        bd_used = inst.currentAZ.get(f'/api/class/fvRsBd.json?'
                                     f'query-target-filter=eq(fvRsBd.tDn,"{inst.bd_dn()}")').json()

        if not bd_used['totalCount']:
            bd.attributes.remove_admin_props()
            bd.delete()

            inst.currentAZ.post(bd.json())

        # Delete Github file
        inst.remove()

        return 200, inst.json()

    def create_drt_instance(self, drenv: str=None):
        apic = APIC(env=(drenv if drenv else self.originAZ.env.DREnv))

        configs = self.generate_config(custom_az=apic.env.Name, drt=True)
        _ = apic.snapshot(descr=f'pre-{self}-drt-inst-creation')
        _ = apic.post(configs.json())

        mapping = self.placeholder_mapping(drt=True)
        _, vlan_info = apic.assign_epg_to_aep(env=apic.env.Name, mapping=mapping)

        return vlan_info

    @classmethod
    def discovery(cls):
        # TODO: Write process to refresh instance settings
        gh = GithubAPI()
        for application in gh.list_dir(cls.GITHUB_PATH):
            for file in gh.list_dir(f'{cls.GITHUB_PATH}/{application}'):
                inst = cls.load(f'{application}/{file}')

                # Check the EPG existence in the current AZ
                r = inst.currentAZ.get(f'/api/mo/{inst.epg_dn()}.json?{FCCO}')
                if int(r.json()['totalCount']):
                    # EPG exists
                    pass
                else:
                    gh.add_file(f'{cls.ISSUES_PATH}/{inst}', message='EPG DN was not found', content=inst.content())
                    continue
                    
                # Check that the EPG still uses the same BD
                r = inst.currentAZ.get(f'/api/mo/{inst.epg_dn()}/rsbd.json')
                if int(r.json()['totalCount']):
                    rsbd = FvRsBd.load(r.json()['imdata'])
                    bd = BD.load(inst.currentAZ.get(f'/api/mo/{rsbd.attributes.tDn}.json?{FCCO}').json()['imdata'])

                    if bd.attributes.dn != inst.bd_dn():
                        # This would mean a new bridge domain is in use, Get the new BD and update the settings
                        if bd.attributes.name != inst.bd_name():
                            inst.bdName = bd.attributes.name

                        # Update BD settings
                        inst.bdSettings = bd.attributes.json()

                    # Check that all networks are found within the BD
                    bd_subnets = bd.get_child_class_iter(Subnet.class_)
                    bd_subnets_set = {_.attributes.ip for _ in bd_subnets}

                    inst_subnets = {_ for _ in inst.networks}
                    if bd_subnets_set.intersection(inst_subnets) != inst_subnets:
                        # One of the networks defined in the instance is missing from the bridge domain
                        gh.add_file(f'{cls.ISSUES_PATH}/{inst}', message='Instance subnet found to be missing from BD',
                                    content=inst.content())
                        continue
                    else:
                        # Everything is good, refresh Subnet settings
                        for subnet in bd_subnets:
                            if subnet.attributes.ip in inst_subnets:
                                u = {subnet.attributes.ip: subnet.attributes.json()}
                                inst.networks.update(u)

                else:
                    # This would mean no BD is assigned to the EPG
                    gh.add_file(f'{cls.ISSUES_PATH}/{inst}', message='No BD assigned to EPG', content=inst.content())

                inst.store()


def update_vlan_spreadsheet():
    envs = json.load(open('data/ACIEnvironments.json'))

    wb = openpyxl.Workbook()

    for env in envs['Environments']:
        if env['Name'] == 'Parallon-Dev':
            continue
        with APIC(env=env['Name']) as apic:
            vlan_data = apic.get_vlan_data()

        vlan_data = json.loads(json.dumps(vlan_data, sort_keys=True))

        xl_data = [['VLAN', 'Consumers', 'AEPs']]

        sheet = wb.create_sheet(env['Name'])

        for key in vlan_data:
            xl_data.append([
                key,
                '\n'.join(vlan_data[key]['Consumers']),
                '\n'.join(vlan_data[key]['AEPs'])
            ])

            for item in xl_data:
                for entry in item:
                    sheet.cell(xl_data.index(item) + 1, item.index(entry) + 1, entry)

        sheet.column_dimensions['B'].width = 96
        sheet.column_dimensions['C'].width = 96

    wb.remove_sheet(wb['Sheet'])

    try:
        smb = SMBConnection(os.getenv('netmgmtuser'), os.getenv('netmgmtpass'), socket.gethostname(),
                            remote_name='corpdpt01.hca.corpad.net', is_direct_tcp=True)
        smb.connect(socket.gethostbyname('corpdpt01.hca.corpad.net'), port=445)

        temp_file = io.BytesIO(save_virtual_workbook(wb))
        smb.storeFile('TELShare',
                      '/Network_Engineering/Network_Design_And_Delivery/Py_ACI_VLANs.xlsx',
                      temp_file)
        del temp_file
        return 200, ['VLAN spreadsheets have been updated.']
    except OperationFailure as e:
        return 500, [f'VLAN spreadsheet Not Updated.  The file is most likely open:  {e}']


def create_dr_env(src_env, dst_env):
    with APIC(env=src_env) as env, APIC(env=dst_env) as drenv:
        if env.snapshot(descr='Auto createDRenv') is False:
            return 500, ['Automated Snapshot Failed.  Task Aborted.']
        if drenv.snapshot(descr='Auto createDrenv') is False:
            return 500, ['Automated Snapshot Failed.  Task Aborted.']

        dr_list = []

        # Collect DR Tags for creating placeholder VLANs
        tags = env.collect_tags('dr')

        for tag in tags:
            dn = tag['tagInst']['attributes']['dn'].split('/')
            dr_list.append(dn[3][4:])

        config = {
            'fvTenant': {
                'attributes': {
                    'name': 'tn-' + env.env.Name,
                    'status': 'created,modified'
                },
                'children': [{
                    'fvCtx': {
                        'attributes': {
                            'name': 'vrf-drtest',
                            'status': 'created,modified'
                            },
                        'children': [{
                            'vzAny': {
                                'attributes': {
                                    'name': ''
                                },
                                'children': [{
                                    'vzRsAnyToProv': {
                                        'attributes': {
                                            'tnVzBrCPName': 'c-drtest-advertise'
                                        }
                                    }
                                }
                                ]
                            }
                        }]
                        }
                    }, {
                    'vzBrCP': {
                        'attributes': {
                            'name': 'c-drtest-advertise',
                            'scope': 'global'
                        },
                        'children': [{
                            'vzSubj': {
                                'attributes': {
                                    'name': 's-drtest-advertise',
                                    'revFltPorts': 'yes'
                                },
                                'children': [{
                                    'vzRsSubjFiltAtt': {
                                        'attributes': {
                                            'tnVzFilterName': 'default'
                                        }
                                    }
                                }]
                            }
                        }]
                    }
                }, {
                    'vzBrCP': {
                        'attributes': {
                            'name': 'c-hca-advertise',
                            'scope': 'global'
                        },
                        'children': [{
                            'vzSubj': {
                                'attributes': {
                                    'name': 's-hca-advertise',
                                    'revFltPorts': 'yes'
                                },
                                'children': [{
                                    'vzRsSubjFiltAtt': {
                                        'attributes': {
                                            'tnVzFilterName': 'default'
                                        }
                                    }
                                }]
                            }
                        }]
                    }
                }
                ]
            }
        }
        admz_config = {
            'fvTenant': {
                'attributes': {
                    'name': 'tn-' + env.env.Name + '-ADMZ',
                    'status': 'created,modified'
                },
                'children': []
            }
        }
        aep_placeholders = {
            'infraAttEntityP': {
                'attributes': {
                    'name': 'aep-Placeholders',
                    'status': 'created,modified'
                },
                'children': [{
                    'infraRsDomP': {
                        'attributes': {
                            'tDn': 'uni/phys-' + drenv.env.PhysicalDomain,
                            'status': 'created,modified'
                            }
                        }
                    }
                ]
            }
        }

        # Collect and add VRFs from the source environment to the configuration
        vrfs = env.collect_vrfs(tn='tn-HCA')
        admz_vrfs = env.collect_vrfs(tn='tn-ADMZ')

        for vrf in vrfs:
            del vrf['fvCtx']['attributes']['dn']
            vrf['fvCtx']['attributes']['status'] = 'created,modified'
            vrf['fvCtx']['children'] = [{
                'vzAny': {
                    'attributes': {
                        'name': ''
                    },
                    'children': [{
                        'vzRsAnyToProv': {
                            'attributes': {
                                'tnVzBrCPName': 'c-hca-advertise'
                            }
                        }
                    }]
                }
            }]
            config['fvTenant']['children'].append(vrf)

        for vrf in admz_vrfs:
            del vrf['fvCtx']['attributes']['dn']
            vrf['fvCtx']['attributes']['status'] = 'created,modified'
            admz_config['fvTenant']['children'].append(vrf)

        # Collect and add bridge domains from the source environment to the configuration
        bds = env.collect_bds(tn='tn-HCA')
        admz_bds = env.collect_bds(tn='tn-ADMZ')

        for bd in bds:
            del bd['fvBD']['attributes']['dn']
            bd['fvBD']['attributes']['status'] = 'created,modified'
            for child in bd['fvBD']['children']:
                if 'fvRsCtx' in child.keys():
                    child['fvRsCtx']['attributes']['tnFvCtxName'] = 'vrf-drtest'
                elif 'fvRsBDToOut' in child.keys():
                    bd['fvBD']['children'].remove(child)
                elif 'fvSubnet' in child.keys():
                    child['fvSubnet']['attributes']['scope'] = 'public,shared'

            config['fvTenant']['children'].append(bd)

        for bd in admz_bds:
            del bd['fvBD']['attributes']['dn']
            bd['fvBD']['attributes']['status'] = 'created,modified'
            admz_config['fvTenant']['children'].append(bd)

        # Collect and add application profiles and from the source environment to the configuration
        aps = env.collect_aps(tn='tn-HCA')
        admz_aps = env.collect_aps(tn='tn-ADMZ')

        for ap in aps:
            del ap['fvAp']['attributes']['dn']
            ap['fvAp']['attributes']['status'] = 'created,modified'

            if ap['fvAp'].get('children'):
                for epg in ap['fvAp']['children']:
                    if 'fvAEPg' in epg.keys():
                        epg['fvAEPg']['attributes']['status'] = 'created,modified'
                        epg['fvAEPg']['children'] = [next(x for x in epg['fvAEPg']['children'] if 'fvRsBd' in x.keys())]
                        epg['fvAEPg']['children'].append({
                            'fvRsDomAtt': {
                                'attributes': {
                                    'tDn': 'uni/phys-' + drenv.env.PhysicalDomain
                                }
                            }
                        })
                        if epg['fvAEPg']['attributes']['name'] in dr_list:
                            epg['fvAEPg']['children'].append({
                                'fvRsCons': {
                                    'attributes': {
                                        'tnVzBrCPName': 'c-drtest-advertise'
                                    }
                                }
                            })
                config['fvTenant']['children'].append(ap)

        for ap in admz_aps:
            del ap['fvAp']['attributes']['dn']
            ap['fvAp']['attributes']['status'] = 'created,modified'

            for epg in ap['fvAp']['children']:
                if 'fvAEPg' in epg.keys():
                    epg['fvAEPg']['attributes']['status'] = 'created,modified'
                    epg['fvAEPg']['children'] = [next(x for x in epg['fvAEPg']['children'] if 'fvRsBd' in x.keys())]
                    epg['fvAEPg']['children'].append({
                        'fvRsDomAtt': {
                            'attributes': {
                                'tDn': 'uni/phys-' + drenv.env.PhysicalDomain
                            }
                        }
                    })
            admz_config['fvTenant']['children'].append(ap)

        # Create snapshot prior to making any changes to the DR fabric
        drenv.snapshot('pre-DR Environment Changes - Automated')

        # If DR environment already exists, delete bridge domains and application profiles to recreate with updated
        # configuration
        if drenv.exists(fvTenant='tn-{}'.format(env.env.Name)) is True:
            dr_aps = drenv.collect_aps(tn='tn-{}'.format(env.env.Name))
            dr_bds = drenv.collect_bds(tn='tn-{}'.format(env.env.Name))

            dr_del = {
                'fvTenant': {
                    'attributes': {
                        'dn': 'uni/tn-tn-' + env.env.Name
                    },
                    'children': []
                }
            }

            for ap in dr_aps:
                ap['fvAp']['attributes']['status'] = 'deleted'
                dr_del['fvTenant']['children'].append(ap)
            for bd in dr_bds:
                bd['fvBD']['attributes']['status'] = 'deleted'
                dr_del['fvTenant']['children'].append(bd)

            # print(json.dumps(dr_del))
            if drenv.post(configuration=dr_del) is False:
                return 500, ['DR Environment exists: Failed to Delete APs and BDs']

        if drenv.post(configuration=config) is False:
            return 500, ['Failed to load tn-{}'.format(env.env.Name)]
        if drenv.post(configuration=admz_config) is False:
            return 500, ['Failed to load tn-{}-ADMZ'.format(env.env.Name)]

        # Check for existence of aep-Placeholders.  If it doesn't exist, create it.
        if drenv.exists(infraAttEntityP='aep-Placeholders') is False:
            if drenv.post(configuration=aep_placeholders) is False:
                return 500, ['aep-Placeholders creation failed.']

        mappings = []

        for tag in tags:
            map_info = tag['tagInst']['attributes']['dn'].split('/')
            if str(map_info[3]).startswith('epg-'):
                if 'ADMZ' in map_info[1]:
                    map_info[1] = f'tn-tn-{env.env.Name}-ADMZ'
                else:
                    map_info[1] = f'tn-tn-{env.env.Name}'

                # Check to see if the DR EPG is already assigned a VLAN
                if drenv.get(f'/api/mo/uni/infra/attentp-aep-Placeholders/gen-default/'
                             f'rsfuncToEpg-[{"/".join(map_info[:-1])}].json').json()['imdata']:
                    continue
                else:
                    aep_map = {
                                'Tenant': map_info[1][3:],
                                'AP': map_info[2][3:],
                                'EPG': map_info[3][4:],
                                'AEP': 'aep-Placeholders'
                            }
                    mappings.append(aep_map)
            else:
                continue

        if mappings:
            for mapping in mappings:
                status, data = APIC.assign_epg_to_aep(drenv.env.Name, mapping)
                if 200 <= status < 300:
                    pass
                else:
                    logging.warning(f'Mapping failed: {mapping}')

        logging.info(f'DR Environment creation completed in {drenv.env.Name}')
        requests.get('https://pyapis.ocp.app.medcity.net/apis/aci/updateVlanSpreadsheets', verify=False)

        return 200, [f'DR Environment creation completed']


# def tag_epgs_v2(env, epgs: list):
#     apic = APIC(env)
#
#     response_data = {}
#
#     for epg in epgs:
#         epg_resp = apic.collect_epgs(epg=epg['epg_name'])
#
#         if epg_resp == 'EPG does not exist':
#             response_data[epg['epg_name']] = epg_resp
#             continue
#
#         tag = GenericClass('tagInst')
#         tag.attributes.dn = f'{epg_resp["fvAEPg"]["attributes"]["dn"]}/tag-{epg["tag"]}'
#         tag.attributes.status = 'created'
#         r = apic.post(configuration=tag.json())
#
#         response_data[epg['epg_name']] = r.reason
#
#     apic.logout()
#
#     return 200, response_data


def migrate_to_admz(env: dict or str, epg: str, subnet: str, next_hop: str, fw_vlan: int or str, tenant: str='tn-HCA'):
    """
    :param env: str or dict
    :param epg: str
    :param subnet: str
    :param next_hop: str
    :param fw_vlan: int or str
    :param tenant: str='tn-HCA'
    :return:

    This function migrates an EPG and Subnet from a specified tenant into the ADMZ tenant.
    """
    def send_config(config: dict):
        r = apic.post(configuration=config)
        response['Configuration Results'].append({f'{r.status_code} | {r.reason}': config,
                                                  'Response Body': r.json()})

    with APIC(env=env) as apic:
        response = {
            'Configuration Results': [],
            'Deleted Configurations': [],
            'New Configurations': [],
            'APIC Snapshot': apic.snapshot(f'Pre-{epg} Migration')
        }

        # Get the specified Subnet
        ip = re.search(r'[\d.]+', subnet).group()

        try:
            bool(IPv4Address(ip))
            bool(IPv4Address(next_hop))
        except AddressValueError:
            return 400, [f'IP Address is not valid: {ip}']

        subnet = apic.collect_subnets(ip=re.search(r'[\d.]+', subnet).group())
        s_bd_name = re.search(r'/BD-([^/]+)', subnet['fvSubnet']['attributes']['dn']).group(1)
        prefix = IPv4Network(subnet['fvSubnet']['attributes']['ip'], strict=False)

        response['Deleted Configurations'].append(subnet)
        subnet['fvSubnet']['attributes']['status'] = 'deleted'

        # Get the specified EPG
        epg = apic.collect_epgs(tn=tenant, epg=epg)

        # Copy the EPG so we can delete it later
        old_epg = deepcopy(epg)
        response['Deleted Configurations'].append(old_epg)
        old_epg['fvAEPg']['attributes']['status'] = 'deleted'

        # Confirm a BD is assigned to the EPG
        if 'fvRsBd' not in list((key for child in epg['fvAEPg']['children'] for key in child.keys())):
            return 400, {'Exception Found': 'Specified EPG does not have a bridge domain assigned', 'EPG': epg}

        # Get the name of the BD used by the EPG and compare it to the BD found to be associated with the subnet
        for child in epg['fvAEPg']['children']:
            if 'fvRsBd' in child.keys():
                if not s_bd_name == child['fvRsBd']['attributes']['tnFvBDName']:
                    return 400, ['The Subnet you specified in not a member of the Bridge Domain used by the EPG you '
                                 'specified']

        epg['fvAEPg']['attributes']['status'] = 'created'

        bd_name = epg['fvAEPg']['attributes']['name'].replace('epg-', 'bd-')
        bd = apic.new_bd(name=bd_name, vrf='vrf-admz', l2=True)
        bd['fvBD']['attributes']['status'] = 'created'

        # Preserve any static paths that may exist
        epg['fvAEPg']['children'] = list(child for child in epg['fvAEPg']['children'] if 'fvRsPathAtt' in child.keys())
        # Append Physical Domain to EPG children
        epg['fvAEPg']['children'].append({'fvRsDomAtt': {'attributes': {'tDn': f'uni/phys-{apic.env.PhysicalDomain}'}}})
        # Append Bridge Domain to EPG children
        epg['fvAEPg']['children'].append({'fvRsBd': {'attributes': {'tnFvBDName': bd_name}}})

        # ap_name = re.search(r'ap-[-\w]+', epg['fvAEPg']['attributes']['dn']).group().replace('ap-ap', 'ap')
        ap_name = re.search(r'/ap-([^/]+)', epg['fvAEPg']['attributes']['dn']).group(1)
        del epg['fvAEPg']['attributes']['dn']
        ap = apic.collect_aps(tn=tenant, ap=ap_name)
        del ap['fvAp']['attributes']['dn']
        ap['fvAp']['attributes']['status'] = 'created,modified'
        epg['fvAEPg']['attributes']['status'] = 'created'
        ap['fvAp']['children'] = [epg]

        admz = {'fvTenant': {'attributes': {'name': 'tn-ADMZ', 'status': 'modified'}, 'children': [bd, ap]}}

        response['New Configurations'].append(admz)

        # Create new EPG and BD in ADMZ Tenant
        send_config(admz)

        # Get static route template for ADMZ
        routes = apic.class_dn_search('ipRouteP', f'{tenant}*out-{apic.env.L3OutADMZ}', get_children=True)

        if isinstance(routes, dict):
            routes = [routes]

        static_route_1 = deepcopy(routes[0])
        static_route_2 = deepcopy(routes[0])

        for key in ['aggregate', 'annotation', 'name', 'nameAlias', 'pref', 'rtCtrl']:
            try:
                del static_route_1['ipRouteP']['attributes'][key]
                del static_route_2['ipRouteP']['attributes'][key]
            except KeyError:
                pass

        static_route_1['ipRouteP']['attributes']['ip'] = prefix.with_prefixlen
        static_route_2['ipRouteP']['attributes']['ip'] = prefix.with_prefixlen
        static_route_1['ipRouteP']['attributes']['status'] = 'created'
        static_route_2['ipRouteP']['attributes']['status'] = 'created'
        static_route_1['ipRouteP']['attributes']['dn'] = re.sub(r'node-\d{3}', 'node-101',
                                                                static_route_1['ipRouteP']['attributes']['dn'])
        static_route_2['ipRouteP']['attributes']['dn'] = re.sub(r'node-\d{3}', 'node-102',
                                                                static_route_2['ipRouteP']['attributes']['dn'])
        static_route_1['ipRouteP']['attributes']['dn'] = re.sub(r'\[[.\d/]+]', f'[{prefix.with_prefixlen}]',
                                                                static_route_1['ipRouteP']['attributes']['dn'])
        static_route_2['ipRouteP']['attributes']['dn'] = re.sub(r'\[[.\d/]+]', f'[{prefix.with_prefixlen}]',
                                                                static_route_2['ipRouteP']['attributes']['dn'])

        static_route_1['ipRouteP']['children'] = [{'ipNexthopP': {'attributes': {'nhAddr': next_hop}}}]
        static_route_2['ipRouteP']['children'] = [{'ipNexthopP': {'attributes': {'nhAddr': next_hop}}}]

        # Create subnet under ADMZ external EPG
        admz_subnet = {
            'l3extSubnet': {
                'attributes': {
                    'dn': f'uni/tn-{tenant}/out-{apic.env.L3OutADMZ}/instP-{apic.env.L3OutADMZ_EPG}/'
                          f'extsubnet-[{prefix.with_prefixlen}]',
                    'ip': prefix.with_prefixlen,
                    'scope': 'import-security'
                }
            }
        }

        # Create subnet under Core external EPG
        core_subnet = {
            'l3extSubnet': {
                'attributes': {
                    'dn': f'uni/tn-{tenant}/out-{apic.env.L3OutCore}/instP-{apic.env.L3OutCore_EPG}/'
                          f'extsubnet-[{prefix.with_prefixlen}]',
                    'ip': prefix.with_prefixlen,
                    'scope': 'export-rtctrl'
                }
            }
        }

        response['New Configurations'].append(static_route_1)
        response['New Configurations'].append(static_route_2)
        response['New Configurations'].append(core_subnet)
        response['New Configurations'].append(admz_subnet)

        # Send static routes and external EPG subnets to APIC then delete old EPG and Subnet
        for configuration in [subnet, static_route_1, static_route_2, admz_subnet, core_subnet, old_epg]:
            send_config(configuration)
            time.sleep(0.1)

        epg_name = epg['fvAEPg']['attributes']['name']

        # Replace with APIC.reassign_encap() / Construct new_epg_dn from tenant, ap, and epg objects
        new_epg_dn = 'uni/tn-%s/ap-%s/epg-%s' % (admz['fvTenant']['attributes']['name'],
                                                 ap['fvAp']['attributes']['name'], epg['fvAEPg']['attributes']['name'])
        apic.reassign_encap(old_epg_dn=old_epg['fvAEPg']['attributes']['dn'], new_epg_dn=new_epg_dn)
        # aep_assignments = apic.class_dn_search('infraRsFuncToEpg', f'{tenant}*{epg_name}', config_only=True)

        # def process_aep_assignments(x):
        #     for item in x:
        #         del item['infraRsFuncToEpg']['attributes']['tDn']
        #
        #         # Delete old mapping
        #         item['infraRsFuncToEpg']['attributes']['status'] = 'deleted'
        #         response['Deleted Configurations'].append(deepcopy(item))
        #         send_config(item)
        #
        #         # Create new mapping
        #         item['infraRsFuncToEpg']['attributes']['dn'] = item['infraRsFuncToEpg']['attributes']['dn'].replace(
        #             tenant, 'tn-ADMZ')
        #         item['infraRsFuncToEpg']['attributes']['status'] = 'created'
        #         response['New Configurations'].append(deepcopy(item))
        #         send_config(item)

        # if isinstance(aep_assignments, dict):
        #     process_aep_assignments([aep_assignments])
        # elif isinstance(aep_assignments, list):
        #     process_aep_assignments(aep_assignments)

        # Create EPG attachment to Firewall AEP
        if apic.env.FirewallAEP != '':
            fw_aep = {
                'infraRsFuncToEpg': {
                    'attributes': {
                        'dn': f'uni/infra/attentp-{apic.env.FirewallAEP}/gen-default/rsfuncToEpg-'
                              f'[uni/tn-tn-ADMZ/ap-{ap_name}/epg-{epg_name}]',
                        'encap': f'vlan-{fw_vlan}',
                        'mode': 'regular',
                        'status': 'created,modified'
                    }
                }
            }
            response['New Configurations'].append(fw_aep)
            send_config(fw_aep)

        # Logic to verify the Bridge Domain is not bound to any other EPGs
        if int(json.loads(apic.get(f'/api/class/fvRsBd.json?query-target-filter=and(wcard(fvRsBd.dn,"{tenant}"),'
                                   f'wcard(fvRsBd.tDn,"{s_bd_name}"))').text)['totalCount']) == 0:
            s_bd = apic.collect_bds(tn=tenant, bd=s_bd_name)
            s_bd['fvBD']['attributes']['status'] = 'deleted'

            response['Deleted Configurations'].append(s_bd)

            # Delete BD that has no binding to anything.  No fvRsBds
            send_config(s_bd)

        # send_config(admz_subnet)

    return 200, response


def assign_epg_to_aep(env, mapping: dict):
    assert 'AEP' and 'Tenant' and 'EPG' and 'AP' in mapping.keys()

    with APIC(env=env) as apic:
        aep = mapping['AEP']
        tenant = mapping['Tenant']
        ap = mapping['AP']
        epg = mapping['EPG']

        epg_data = EPG.load(apic.collect_epgs(tenant, epg))
        if ap not in epg_data.attributes.dn:
            return 500, ['EPG lookup error']

        epg_vlan_data = apic.get_vlan_data(epg=epg)

        if epg_vlan_data == list():
            vlan = apic.get_next_vlan()
        elif len(epg_vlan_data) > 1:
            try:
                vlan = next(key for vlan in epg_vlan_data for key in vlan if 'aep-Placeholders' in vlan[key]['AEPs'])
            except StopIteration:
                return 500, ['API could not determine VLAN to use as multiple are in use and designation is not '
                             'defined on aep-Placholders']
        else:
            vlan = next(key for key in epg_vlan_data[0])

        aep_c = AEP()
        aep_c.attributes.name = aep
        aep_c.attributes.status = 'modified'
        aep_c.add_epg(epg_dn=epg_data.attributes.dn, encap=vlan)

        r = apic.post(configuration=aep_c.json(), uri=aep_c.post_uri)

        if not r.ok:
            return r.status_code, json.loads(r.text)

    return 200, {'EPG': epg, 'VLAN': vlan}


def find_subnet(req_data):
    envs = json.load(open('data/ACIEnvironments.json', 'r'))
    ip = req_data['ip']

    resp = ''
    env = envs['Environments'][0]

    for env in envs['Environments']:
        with APIC(env=env['Name']) as apic:
            resp = apic.collect_subnets(ip=ip)
        if resp.__class__ is not str:
            break

    if isinstance(resp, dict):
        if 'fvSubnet' in resp.keys():
            bd = resp['fvSubnet']['attributes']['dn'][resp['fvSubnet']['attributes']['dn'].index('BD-') + 3:
                                                      resp['fvSubnet']['attributes']['dn'].index('/subnet')]
            network = IPv4Network(resp['fvSubnet']['attributes']['ip'], strict=False)
            subnet = {
                'AvailabilityZone': env['Name'],
                'BridgeDomain': bd,
                'Subnet': network.with_prefixlen
            }
            return subnet
        elif 'l3extSubnet' in resp.keys():
            l3out = resp['l3extSubnet']['attributes']['dn'][resp['l3extSubnet']['attributes']['dn'].index('out-') + 4:
                                                            resp['l3extSubnet']['attributes']['dn'].index('/instP-')]
            epg = resp['l3extSubnet']['attributes']['dn'][resp['l3extSubnet']['attributes']['dn'].index('instP-') + 6:
                                                          resp['l3extSubnet']['attributes']['dn'].index('/extsubnet-')]
            subnet = {
                'AvailabilityZone': env['Name'],
                'L3Out': l3out,
                'ExternalEPG': epg,
                'Network': resp['l3extSubnet']['attributes']['ip']
            }
            return subnet
        else:
            return str(resp)
    else:
        return resp


def fabric_inventory():
    try:
        envs = json.load(open('data/ACIEnvironments.json', 'r'))
        data = []

        for env in envs['Environments']:
            with APIC(env=env['Name']) as apic:
                invs = apic.collect_inventory()

            for inv in invs:
                if 'error' not in inv.keys():
                    data.append(inv)

        fabrics = list(set((x['topSystem']['attributes']['fabricDomain'] for x in data)))
        fabrics = {each: [] for each in fabrics}

        for inv in data:
            assignment = inv['topSystem']['attributes']['fabricDomain']
            fabrics[assignment].append(inv['topSystem']['attributes'])

        wb = openpyxl.Workbook()

        for fabric in fabrics:
            xl_data = [['SwitchID', 'Serial Number', 'Name', 'OOB IP']]
            for node in fabrics[fabric]:
                xl_data.append([
                    int(re.search('[0-9]+', re.search('node-[0-9]+', node['dn']).group()).group()),
                    node['serial'],
                    node['name'],
                    node['oobMgmtAddr']
                ])

            sheet = wb.create_sheet(fabric)

            for item in xl_data:
                for entry in item:
                    sheet.cell(xl_data.index(item) + 1, item.index(entry) + 1, entry)

        wb.remove_sheet(wb['Sheet'])

        smb = SMBConnection(os.getenv('netmgmtuser'), os.getenv('netmgmtpass'), socket.gethostname(),
                            remote_name='corpdpt01.hca.corpad.net', is_direct_tcp=True)
        smb.connect(socket.gethostbyname('corpdpt01.hca.corpad.net'), port=445)
        temp_file = io.BytesIO(save_virtual_workbook(wb))
        smb.storeFile('TELShare',
                      '/Network_Engineering/ACI/ACI_Serial_Numbers/Py_ACI_Serials.xlsx',
                      temp_file)
        del temp_file
        return 200, fabrics

    except OperationFailure as e:
        return 500, [f'Inventory Not Updated.  The file is most likely open:  {e}']


def configure_interfaces(env: str, req_data: dict):
    responses = []
    with APIC(env=env) as apic:
        for aep in req_data:
            if not aep['PortTemplateName'].startswith('aep-'):
                responses.append({
                    aep['PortTemplateName']: f'Not a valid Port Template.  '
                                             f'No action taken for {aep["PortTemplateName"]}'
                })
                continue

            r, configuration = apic.interface_configuration(aep_name=aep['PortTemplateName'],
                                                            infra_info=aep['Configurations'])

            if not r.ok:
                responses.append({
                    r.status_code: {r.text: configuration}
                })
            else:
                responses.append({
                    r.status_code: configuration
                })

    return 200, responses


def create_new_epg(env: str, req_data: dict):
    with APIC(env=env) as apic:
        big = BIG()

        ap_name = req_data['AppProfileName']
        epg_name = req_data['EPGName']
        bd_name = req_data['BridgeDomainName']
        description = req_data['Description']
        no_of_ips = req_data['NumIPsReqd']

        if not ap_name.startswith('ap-'):
            ap_name = f'ap-{ap_name}'
        if not epg_name.startswith('epg-'):
            epg_name = f'epg-{epg_name}'
        if not bd_name.startswith('bd-'):
            bd_name = f'bd-{bd_name}'

        # Check if EPG/BD exists prior to assigning a network
        if apic.dn_exists(fvAEPg=f'uni/tn-{apic.env.Tenant}/ap-{ap_name}/epg-{epg_name}'):
            return 400, {'message': 'The requested EPG already exists'}

        if apic.bd_exists(tn=apic.env.Tenant, bd_name=bd_name):
            return 400, {'message': 'The requested bridge domain already exists'}

        subnet = big.assign_next_network_from_list(block_list=apic.env.Subnets, no_of_ips=no_of_ips, name=epg_name,
                                                   coid=int(apic.env.COID), asn=int(apic.env.ASN))

        if not subnet:
            return 400, [f'{apic.env.Name} has no available network for the required number of IPs.']

        network = IPv4Network(subnet.properties['CIDR'])
        gateway = network.network_address + 1

        # Define Tenant
        tn = Tenant()
        tn.attributes.name = apic.env.Tenant
        tn.attributes.status = 'modified'

        # Define Application Profile
        ap = AP()
        ap.attributes.name = ap_name
        ap.attributes.status = 'created,modified'

        # Define Endpoint Group
        epg = EPG()
        epg.attributes.name = epg_name
        epg.attributes.descr = description
        epg.attributes.status = 'created'
        epg.domain(name=apic.env.PhysicalDomain)

        bd = BD()
        bd.layer3()
        # Add Subnet to Bridge Domain and set Description on Subnet
        bd.add_subnet(subnet=f'{gateway}/{network.prefixlen}')
        bd.children[0].attributes.descr = description
        # Configure the bridge domain to use the environment VRF
        bd.use_vrf(name=apic.env.VRF)
        # Attach the Core L3Out to the Bridge Domain
        bd.to_l3_out(name=apic.env.L3OutCore)
        bd.attributes.name = bd_name
        bd.attributes.status = 'created,modified'

        # Assign BD to EPG
        epg.assign_bd(bd.attributes.name)

        # Attach Bridge Domain and Application Profile as child objects of Tenant
        tn.children.append(bd)
        tn.children.append(ap)
        # Attach EPG as child object of Application Profile
        ap.children.append(epg)

        # POST the configuration to APIC
        epg_r = apic.post(configuration=tn.json())

        # Check response status.  Abort if not OK.
        if not epg_r.ok:
            return 400, [epg_r.status_code, epg_r.reason, epg_r.text, tn.json()]

        # Get status of aep-Placeholders and assign EPG to it
        aep = AEP()
        aep.attributes.name = 'aep-Placeholders'

        if not apic.exists(infraAttEntityP='aep-Placeholders'):
            aep.attributes.status = 'created'
            dom_p = GenericClass('infraRsDomP')
            dom_p.attributes.tDn = f'uni/phys-{apic.env.PhysicalDomain}'
            aep.children.append(dom_p)
        else:
            aep.attributes.status = 'modified'

        # Get the EPG
        g_epg = EPG.load(apic.collect_epgs(tn=apic.env.Tenant, epg=epg.attributes.name))

        # Get next available VLAN for the EPG
        vlan = apic.get_next_vlan()

        # Update the network in Proteus
        subnet.properties['VLAN'] = vlan

        big.update_object(subnet)
        big.logout()

        # Add the EPG to the AEP with the VLAN returned
        aep.add_epg(epg_dn=g_epg.attributes.dn, encap=vlan)
        # POST the AEP configuration to APIC
        apic.post(configuration=aep.json(), uri=aep.post_uri)

    # Add the instance to Github
    inst = AppInstance(application=ap.attributes.name,
                       name=epg.attributes.name,
                       apName=ap.attributes.name,
                       bdName=bd.attributes.name,
                       epgName=epg.attributes.name,
                       bdSettings=bd.attributes.json(),
                       tenant=next(_ for _ in dir(apic.env)
                                   if apic.env.__getattribute__(_) == tn.attributes.name),
                       networks={n.attributes.ip: n.attributes.json() for n in bd.get_child_class_iter(Subnet.class_)},
                       currentAZ=str(apic))

    inst.store()

    return 200, {'EPG Name': epg.attributes.name, 'Subnet': subnet.properties['CIDR'], 'VLAN': vlan,
                 'AppInstance': inst.json(), 'InstancePath': f'{inst.application}/{inst}'}


def create_custom_epg(env: str, req_data: dict):
    with APIC(env=env) as apic:
        ap_name = req_data['AppProfileName']
        epg_name = req_data['EPGName']
        bd_name = req_data['BridgeDomainName']
        description = req_data['Description']
        subnets = req_data['Subnets']

        # Define Tenant
        tn = Tenant()
        tn.attributes.name = apic.env.Tenant
        tn.attributes.status = 'modified'

        # Define Application Profile
        ap = AP()
        ap.attributes.name = ap_name
        ap.attributes.status = 'created,modified'

        # Define Endpoint Group
        epg = EPG()
        epg.attributes.name = epg_name
        epg.attributes.descr = description
        epg.attributes.status = 'created'
        epg.domain(name=apic.env.PhysicalDomain)

        # Check to see if BD already exists
        bd_exists = apic.bd_exists(tn=apic.env.Tenant, bd_name=bd_name)

        if bool(bd_exists):
            bd = BD.load(bd_exists)
            for subnet in subnets:
                network = IPv4Network(subnet)
                gateway = network.network_address + 1
                bd.add_subnet(subnet=f'{gateway}/{network.prefixlen}', description=description)
            bd.attributes.status = 'modified'
        else:
            bd = BD()
            bd.layer3()
            # Add Subnet to Bridge Domain and set Description on Subnet
            for subnet in subnets:
                network = IPv4Network(subnet)
                gateway = network.network_address + 1
                bd.add_subnet(subnet=f'{gateway}/{network.prefixlen}', description=description)
            # Configure the bridge domain to use the environment VRF
            bd.use_vrf(name=apic.env.VRF)
            # Attach the Core L3Out to the Bridge Domain
            bd.to_l3_out(name=apic.env.L3OutCore)
            bd.attributes.name = bd_name
            bd.attributes.status = 'created,modified'

        # Assign BD to EPG
        epg.assign_bd(bd.attributes.name)

        # Attach Bridge Domain and Application Profile as child objects of Tenant
        tn.children.append(bd)
        tn.children.append(ap)
        # Attach EPG as child object of Application Profile
        ap.children.append(epg)

        # POST the configuration to APIC
        epg_r = apic.post(configuration=tn.json())

        # Check response status.  Abort if not OK.
        if not epg_r.ok:
            return 400, [epg_r.status_code, epg_r.reason, epg_r.json(), tn.json()]

        # Get status of aep-Placeholders and assign EPG to it
        aep = AEP()
        aep.attributes.name = 'aep-Placeholders'

        if not apic.exists(infraAttEntityP='aep-Placeholders'):
            aep.attributes.status = 'created'
            dom_p = GenericClass('infraRsDomP')
            dom_p.attributes.tDn = f'uni/phys-{apic.env.PhysicalDomain}'
            aep.children.append(dom_p)
        else:
            aep.attributes.status = 'modified'

        # Get the EPG
        g_epg = EPG.load(apic.collect_epgs(tn=apic.env.Tenant, epg=epg.attributes.name))

        # Get next available VLAN for the EPG
        vlan = apic.get_next_vlan()

        # Add the EPG to the AEP with the VLAN returned
        aep.add_epg(epg_dn=g_epg.attributes.dn, encap=vlan)
        # POST the AEP configuration to APIC
        apic.post(configuration=aep.json(), uri=aep.post_uri)

    # Add the instance to GitHub
    inst = AppInstance(application=ap.attributes.name,
                       name=epg.attributes.name,
                       apName=ap.attributes.name,
                       bdName=bd.attributes.name,
                       epgName=epg.attributes.name,
                       bdSettings=bd.attributes.json(),
                       tenant=next(_ for _ in dir(apic.env)
                                   if apic.env.__getattribute__(_) == tn.attributes.name),
                       networks={n.attributes.ip: n.attributes.json() for n in bd.get_child_class_iter(Subnet.class_)},
                       currentAZ=str(apic))
    inst.store()

    return 200, {'EPG Name': epg.attributes.name, 'Subnets': subnets, 'VLAN': vlan,
                 'AppInstance': inst.json(), 'InstancePath': f'{inst.application}/{inst}'}


def create_custom_epg_v2(env: str, req_data: dict):
    with APIC(env=env) as apic:
        ap_name = req_data['AppProfileName']
        epg_name = req_data['EPGName']
        bd_name = req_data['BridgeDomainName']
        description = req_data['Description']
        subnets = req_data['Subnets']
        tn_name = req_data['TenantName']
        vrf_name = req_data['VRFName']

        # Define Tenant
        tn = Tenant()
        tn.attributes.name = tn_name
        tn.attributes.status = 'modified'

        # Define Application Profile
        ap = AP()
        ap.attributes.name = ap_name
        ap.attributes.status = 'created,modified'

        # Define Endpoint Group
        epg = EPG()
        epg.attributes.name = epg_name
        epg.attributes.descr = description
        epg.attributes.status = 'created'
        epg.domain(name=apic.env.PhysicalDomain)

        # Check to see if BD already exists
        bd_exists = apic.bd_exists(tn=tn_name, bd_name=bd_name)

        if bool(bd_exists):
            bd = BD.load(bd_exists)
            for subnet in subnets:
                network = IPv4Network(subnet)
                gateway = network.network_address + 1
                bd.add_subnet(subnet=f'{gateway}/{network.prefixlen}', description=description)
            bd.attributes.status = 'modified'
        else:
            bd = BD()
            bd.layer3()
            # Add Subnet to Bridge Domain and set Description on Subnet
            for subnet in subnets:
                network = IPv4Network(subnet)
                gateway = network.network_address + 1
                bd.add_subnet(subnet=f'{gateway}/{network.prefixlen}', description=description)
            # Configure the bridge domain to use the environment VRF
            bd.use_vrf(name=vrf_name)
            # Attach the Core L3Out to the Bridge Domain
            bd.attributes.name = bd_name
            bd.attributes.status = 'created,modified'

        # Assign BD to EPG
        epg.assign_bd(bd.attributes.name)

        # Attach Bridge Domain and Application Profile as child objects of Tenant
        tn.children.append(bd)
        tn.children.append(ap)
        # Attach EPG as child object of Application Profile
        ap.children.append(epg)

        # POST the configuration to APIC
        epg_r = apic.post(configuration=tn.json())

        # Check response status.  Abort if not OK.
        if not epg_r.ok:
            return 400, [epg_r.status_code, epg_r.reason, epg_r.json(), tn.json()]

        # Get status of aep-Placeholders and assign EPG to it
        aep = AEP()
        aep.attributes.name = 'aep-Placeholders'

        if not apic.exists(infraAttEntityP='aep-Placeholders'):
            aep.attributes.status = 'created'
            dom_p = GenericClass('infraRsDomP')
            dom_p.attributes.tDn = f'uni/phys-{apic.env.PhysicalDomain}'
            aep.children.append(dom_p)
        else:
            aep.attributes.status = 'modified'

        # Get the EPG
        g_epg = EPG.load(apic.collect_epgs(tn=tn_name, epg=epg.attributes.name))

        # Get next available VLAN for the EPG
        vlan = apic.get_next_vlan()

        # Add the EPG to the AEP with the VLAN returned
        aep.add_epg(epg_dn=g_epg.attributes.dn, encap=vlan)
        # POST the AEP configuration to APIC
        apic.post(configuration=aep.json(), uri=aep.post_uri)

        # Only document the instance IF the EPG is being created in a production Tenant
        if tn.attributes.name in [apic.env.__getattribute__('ADMZTenant'), apic.env.__getattribute__('Tenant')]:
            inst = AppInstance(application=ap.attributes.name,
                               name=epg.attributes.name,
                               apName=ap.attributes.name,
                               bdName=bd.attributes.name,
                               epgName=epg.attributes.name,
                               bdSettings=bd.attributes.json(),
                               tenant=next(_ for _ in dir(apic.env)
                                           if apic.env.__getattribute__(_) == tn.attributes.name),
                               networks={n.attributes.ip: n.attributes.json() for n in
                                         bd.get_child_class_iter(Subnet.class_)},
                               currentAZ=str(apic))
            inst.store()

    return 200, {'EPG Name': epg.attributes.name, 'Subnets': subnets, 'VLAN': vlan,
                 'Message': f'Please add the appropriate L3Out profile to {bd_name}',
                 'AppInstance': inst.json(), 'InstancePath': f'{inst.application}/{inst}'}


def create_new_admz_epg(env: str, req_data: dict):
    with APIC(env=env) as apic:
        big = BIG()

        ap_name = req_data['AppProfileName']
        epg_name = req_data['EPGName']
        description = req_data['Description']
        no_of_ips = req_data['NumIPsReqd']

        subnet = big.assign_next_network_from_list(block_list=apic.env.ADMZSubnets, no_of_ips=no_of_ips, name=epg_name,
                                                   coid=int(apic.env.COID), asn=int(apic.env.ASN))

        if not subnet:
            big.logout()
            return 400, [f'{apic.env.Name} has no available network for the required number of IPs.']

        big.logout()

        # network = IPv4Network(subnet.properties['CIDR'])
        # gateway = network.network_address + 1

        # Define Tenant
        tn = Tenant()
        tn.attributes.name = apic.env.ADMZTenant
        tn.attributes.status = 'modified'

        # Define Application Profile
        ap = AP()
        ap.attributes.name = ap_name
        ap.attributes.status = 'created,modified'

        # Define Endpoint Group
        epg = EPG()
        epg.attributes.name = epg_name
        epg.attributes.descr = description
        epg.attributes.status = 'created'
        epg.domain(name=apic.env.PhysicalDomain)

        # Define Bridge Domain
        bd = BD()
        bd.layer2()
        # Configure the bridge domain to use the environment VRF
        bd.use_vrf(name=apic.env.ADMZVRF)
        bd.attributes.name = epg_name.replace('epg-', 'bd-')
        bd.attributes.status = 'created,modified'

        # Assign BD to EPG
        epg.assign_bd(bd.attributes.name)

        # Attach Bridge Domain and Application Profile as child objects of Tenant
        tn.children.append(bd)
        tn.children.append(ap)
        # Attach EPG as child object of Application Profile
        ap.children.append(epg)

        # POST the configuration to APIC
        epg_r = apic.post(configuration=tn.json())

        # Check response status.  Abort if not OK.
        if not epg_r.ok:
            return 400, [epg_r.status_code, epg_r.reason, epg_r.text, tn.json()]

        # Get status of aep-Placeholders and assign EPG to it
        aep = AEP()
        aep.attributes.name = 'aep-Placeholders'

        if not apic.exists(infraAttEntityP='aep-Placeholders'):
            aep.attributes.status = 'created'
            dom_p = GenericClass('infraRsDomP')
            dom_p.attributes.tDn = f'uni/phys-{apic.env.PhysicalDomain}'
            aep.children.append(dom_p)
        else:
            aep.attributes.status = 'modified'

        # Get the EPG
        g_epg = EPG.load(apic.collect_epgs(tn=apic.env.ADMZTenant, epg=epg.attributes.name))

        # Get next available VLAN for the EPG
        vlan = apic.get_next_vlan()
        # Add the EPG to the AEP with the VLAN returned
        aep.add_epg(epg_dn=g_epg.attributes.dn, encap=vlan)
        # POST the AEP configuration to APIC
        apic.post(configuration=aep.json(), uri=aep.post_uri)

        fw_aep = AEP.load(apic.collect_aeps(aep=apic.env.FirewallAEP))
        fw_aep.add_epg(epg_dn=g_epg.attributes.dn, encap=vlan)
        apic.post(configuration=fw_aep.json(), uri=aep.post_uri)

    # TODO: Add a process to configure the VLAN interface on the firewall:  Handle this in main.py

    # TODO: Add processes to update CMDB

    return 200, {'EPG Name': epg.attributes.name, 'Subnet': subnet.properties['CIDR'], 'VLAN': vlan}


def add_oob_leaf(env: str, rack: str, serial: str, pod: int=1):
    response = {}

    with APIC(env=env) as apic:
        # Assert that maintenance policies and groups exist for assignment
        apic.verify_maintenance_policies_and_groups()

        try:
            node_id = apic.get_next_oob_node(pod=pod)
        except LeafLimitError:
            return 400, {'message': 'OOB leaf limit reached for the specified pod'}
        except AssertionError:
            return 400, {'message': 'Pod ID out of range'}

        # Register the leaf (Node, Serial and rack validations handled within APIC.register_leaf())
        node_policy, r = apic.register_leaf(rack=rack, serial=serial, node_id=node_id)
        if node_policy == 400:
            return node_policy, r

        if not r.ok:
            return r.status_code, json.loads(r.text)

        # Pause before attempting OOB assignment
        time.sleep(2)

        # Get and Create OOB Address for new node: Address assignment handled within APIC.assign_oob_to_leaf())
        oob_address, r = apic.assign_oob_to_leaf(node_id=node_policy.attributes.nodeId,
                                                 node_name=node_policy.attributes.name)
        response['OOB Assignment Operation'] = {f'{r.status_code} | {r.reason}': {'Request Body': oob_address.json(),
                                                                                  'Response Body': json.loads(r.text)}}

        # Assign node to Staging Maintenance Group
        config, r = apic.assign_leaf_to_maint_group(apic.env.StagingMaintenanceGroup,
                                                    node=node_policy.attributes.nodeId)
        response['Firmware Staging'] = {f'{r.status_code} | {r.reason}': {'Request Body': config.json(),
                                                                          'Response Body': json.loads(r.text)}}

        # Schedule Maintenance Group move 12 hours from now
        configs = []

        new_block = FabricNodeBlock.load(config.json())
        new_block.attributes.dn = 'uni/fabric/maintgrp-%s/nodeblk-blk%s-%s' % (apic.env.OOBMaintenanceGroup,
                                                                               node_id, node_id)

        new_block.create()
        config.delete()
        configs.append({'configuration': config.json(), 'uri': '/api/mo/uni.json'})
        configs.append({'configuration': new_block.json(), 'uri': '/api/mo/uni.json'})
        ACIJob.create_aci_job(job_name=f'{apic.env.Name}-{node_id}_maintgrp_move', environment=apic.env.Name,
                              delay_in_seconds=43200, configs=configs, func=add_oob_leaf.__name__)

        # Add node to OOB Leaf switch profile
        switch_profile, r = apic.add_leaf_to_oob_profile(node_policy.attributes.nodeId)
        response['Leaf Profile Operation'] = {f'{r.status_code} | {r.reason}': {'Request Body': switch_profile.json(),
                                                                                'Response Body': json.loads(r.text)}}

    return 200, response


def add_new_leaf_pair(env: str, rack1: str, serial1: str, rack2: str, serial2: str, pod: int=1):
    response = {}

    # Validate both rack numbers prior to implementing any configurations
    if not re.match(r'\w\d\d', rack1) or not re.match(r'\w\d\d', rack2):
        return 400, ['Invalid Rack number format']

    with APIC(env=env) as apic:
        # Assert the existence of maintenance policies and groups
        apic.verify_maintenance_policies_and_groups()

        # Validate both serial numbers prior to implementing any configurations
        if not apic.verify_serial(serial=serial1) or not apic.verify_serial(serial=serial2):
            return 400, ['Serial Not found on fabric']

        # Get Node IDs for new switches
        try:
            node1, node2 = apic.get_next_leaf_pair(pod=pod)
        except LeafLimitError:
            return 400, {'message': 'Leaf limit has been reached for the selected pod'}
        except AssertionError:
            return 400, {'message': 'Pod ID out of range'}

        # Validate both Node IDs prior to implementing any configurations
        if apic.dn_exists(fabricNode=f'node-{node1}') or apic.dn_exists(fabricNode=f'node-{node2}'):
            return 500, ['Internal Error Occurred. Nodes IDs that already exist were selected.  Operation Aborted.']

        # Register the leafs
        node_policy1, r1 = apic.register_leaf(rack=rack1.upper(), serial=serial1.upper(), node_id=node1)
        node_policy2, r2 = apic.register_leaf(rack=rack2.upper(), serial=serial2.upper(), node_id=node2)

        if node_policy1 == 400 or node_policy2 == 400:
            return 400, [r1, r2]

        if not r1.ok and r2.ok:
            return 400, [{r1.status_code: json.loads(r1.text)}, {r2.status_code: json.loads(r2.text)}]

        # Pause before attempting OOB assignment
        time.sleep(2)

        # Configure OOB Addresses on each of the new nodes
        oob_address1, r1 = apic.assign_oob_to_leaf(node_id=node_policy1.attributes.nodeId,
                                                   node_name=node_policy1.attributes.name)
        response[f'{node1} OOB Assignment Operation'] = {f'{r1.status_code} | {r1.reason}':
                                                         {'Request Body': oob_address1.json(),
                                                          'Response Body': json.loads(r1.text)}}
        oob_address2, r2 = apic.assign_oob_to_leaf(node_id=node_policy2.attributes.nodeId,
                                                   node_name=node_policy2.attributes.name)
        response[f'{node2} OOB Assignment Operation'] = {f'{r2.status_code} | {r2.reason}':
                                                         {'Request Body': oob_address2.json(),
                                                          'Response Body': json.loads(r2.text)}}

        config1, r1 = apic.assign_leaf_to_maint_group(apic.env.StagingMaintenanceGroup, node1)
        config2, r2 = apic.assign_leaf_to_maint_group(apic.env.StagingMaintenanceGroup, node2)

        response[f'{node1} Firmware Staging'] = {f'{r1.status_code} | {r1.reason}': {'Request Body': config1.json(),
                                                                                     'Response Body': json.loads(
                                                                                         r1.text)}}
        response[f'{node2} Firmware Staging'] = {f'{r2.status_code} | {r2.reason}': {'Request Body': config2.json(),
                                                                                     'Response Body': json.loads(
                                                                                         r2.text)}}

        # Schedule Maintenance Group move 12 hours from now
        configs = []

        new_block1 = FabricNodeBlock.load(config1.json())
        new_block2 = FabricNodeBlock.load(config2.json())
        config1.delete()
        config2.delete()

        new_block1.attributes.dn = 'uni/fabric/maintgrp-%s/nodeblk-blk%s-%s' % (apic.env.OddsMaintenanceGroup,
                                                                                node_policy1.attributes.nodeId,
                                                                                node_policy1.attributes.nodeId
                                                                                )
        new_block2.attributes.dn = 'uni/fabric/maintgrp-%s/nodeblk-blk%s-%s' % (apic.env.EvensMaintenanceGroup,
                                                                                node_policy2.attributes.nodeId,
                                                                                node_policy2.attributes.nodeId
                                                                                )

        new_block1.create()
        new_block2.create()
        configs.append({'configuration': config1.json(), 'uri': '/api/mo/uni.json'})
        configs.append({'configuration': config2.json(), 'uri': '/api/mo/uni.json'})
        configs.append({'configuration': new_block1.json(), 'uri': '/api/mo/uni.json'})
        configs.append({'configuration': new_block2.json(), 'uri': '/api/mo/uni.json'})
        ACIJob.create_aci_job(job_name=f'{apic.env.Name}-{node1}-{node2}_maintgrp_move', environment=apic.env.Name,
                              delay_in_seconds=43200, configs=configs, func=add_new_leaf_pair.__name__)

        # Create Switch Profile and VPC
        rack_set = {rack1, rack2}

        profile_name = f'{apic.env.Name}-{"-".join(rack_set)}-LF-{node1}-{node2}'

        switch_profile, r = apic.create_leaf_pair_profile(name=profile_name, nodes=[node1, node2])
        response['Leaf Profile Operation'] = {f'{r.status_code} | {r.reason}': {'Request Body': switch_profile.json(),
                                                                                'Response Body': json.loads(r.text)}}
    return 200, response


def get_epg_data(environment: str, epg: str):
    subnets = set()

    with APIC(env=environment) as apic:
        apic_subnets = apic.get('/api/class/fvSubnet.json').json()['imdata']
        apic_subnets = list(IPv4Network(subnet['fvSubnet']['attributes']['ip'], strict=False)
                            for subnet in apic_subnets)

        endpoints = apic.get(f'/api/class/fvCEp.json?query-target-filter=wcard(fvCEp.dn,"{epg}")').json()['imdata']
        if endpoints:
            epg_dn = re.search(r'(.*)/cep', list(endpoints)[0]['fvCEp']['attributes']['dn']).group(1)
            epg = APICObject.load(apic.get('/api/mo/%s.json?%s' % (epg_dn, FULL_CONFIG)).json()['imdata'][0])
        else:
            epg = apic.collect_epgs(epg=epg)
            epg = APICObject.load(epg)

        endpoint_count = len(endpoints)
        vlans = set(int(endpoint['fvCEp']['attributes']['encap'][5:]) for endpoint in endpoints)
        endpoints = set(endpoint['fvCEp']['attributes']['ip'] for endpoint in endpoints
                        if endpoint['fvCEp']['attributes']['ip'] != '0.0.0.0')

    bd = next(child for child in epg.children if child.class_ == 'fvRsBd')

    for endpoint in endpoints:
        for subnet in apic_subnets:
            if subnet.overlaps(IPv4Network(endpoint)):
                subnets.add(subnet.with_prefixlen)

    response = {
        'epg_name': epg.attributes.name,
        'bridge_domain': bd.attributes.tnFvBDName,
        'subnets': list(subnets),
        'vlans': list(vlans),
        'epg_dn': epg.attributes.dn,
        'endpoint_count': endpoint_count
    }

    return 200, response


def get_aci_vlan_data(environment, encap: str):
    return_data = []

    with APIC(env=environment) as apic:
        vlan_usage = apic.get_vlan_data(vlan=int(encap))
    vlan_usage = vlan_usage[int(encap)]['Consumers']

    for consumer in vlan_usage:
        epg = consumer.split('/')[-1][4:]
        _, response = get_epg_data(environment=environment, epg=epg)

        return_data.append(response)

    return 200, return_data


def get_aci_subnet_data(environment, ip: str):
    apic = APIC(env=environment)

    subnet = APICObject.load(apic.collect_subnets(ip=ip))
    # TODO: Add ability to process L3Out networks - l3extSubnet

    if subnet.class_ == Subnet.class_:
        network = IPv4Network(subnet.attributes.ip, strict=False)
        eps = [APICObject.load(_) for _ in apic.collect_eps()]
        eps = [_ for _ in eps if network.overlaps(IPv4Network(_.attributes.ip))]

        vlans = list(set(int(re.search(r'\d+', _.attributes.encap).group()) for _ in eps))
        epgs = list(set(re.search(r'(.*)/cep-', _.attributes.dn).group(1) for _ in eps))
        bd = re.search(r'BD-([^/]+)', subnet.attributes.dn).group(1)

        return 200, {'subnet': network.with_prefixlen, 'vlans': vlans, 'consumers': epgs, 'bridge_domain': bd,
                     'endpoint_count': len(eps), 'l3out': None}
    elif subnet.class_ == 'l3extSubnet':
        network = IPv4Network(subnet.attributes.ip, strict=False)
        consumer = re.search(r'(.*)/extsubnet', subnet.attributes.dn).group(1)
        l3out = re.search(r'/out-([^/]+)', subnet.attributes.dn).group(1)

        return 200, {'subnet': network.with_prefixlen, 'vlans': [], 'consumers': [consumer], 'bridge_domain': None,
                     'endpoint_count': None, 'l3out': l3out}

    else:
        return 404, {'error': 'The requested network was not found'}


def fabric_status(env: str, username: str=None, password: str=None):
    timestamp = datetime.now()
    timestamp = datetime.strptime(timestamp.ctime(), '%c')
    start = time.perf_counter()

    with APIC(env=env, username=username, password=password) as apic:
        api = GithubAPI()

        # Collect all IP Routes (ipRoute)
        ip_routes = apic.get('/api/class/ipRoute.json').json()['imdata']
        ip_routes = list((r['ipv4Route']['attributes']['dn'] for r in ip_routes))
        # Collect all OSPF routes (ospfRoute)
        ospf_routes = apic.get('/api/class/ospfRoute.json').json()['imdata']
        ospf_routes = list((o['ospfRoute']['attributes']['dn'] for o in ospf_routes))
        # Collect all routing protocol neighbors (ospfAdjEp)
        ospf_neighbors = apic.get('/api/class/ospfAdjEp.json').json()['imdata']
        ospf_neighbors = list((o['ospfAdjEp']['attributes']['dn'] for o in ospf_neighbors))
        # Collect BGP peers (bgpPeer)
        bgp_peers = apic.get('/api/class/bgpPeer.json').json()['imdata']
        bgp_peers = list(p['bgpPeer']['attributes']['dn'] for p in bgp_peers)
        # Collect all endpoints (fvCEp)
        endpoints = apic.collect_eps()
        endpoints = list((endpoint[key]['attributes']['ip'] for endpoint in endpoints for key in endpoint
                          if endpoint[key]['attributes']['ip'] != '0.0.0.0'))
        # Collect all faults (faultInst)
        minor_faults = apic.get('/api/class/faultInst.json?query-target-filter='
                                'eq(faultInst.severity,"minor")').json()['imdata']
        minor_faults = list(({'code': f['faultInst']['attributes']['code'],
                              'cause': f['faultInst']['attributes']['cause'],
                              'severity': f['faultInst']['attributes']['severity'],
                              'descr': f['faultInst']['attributes']['descr'],
                              'dn': f['faultInst']['attributes']['dn']}
                             for f in minor_faults))
        warning_faults = apic.get('/api/class/faultInst.json?query-target-filter='
                                  'eq(faultInst.severity,"warning")').json()['imdata']
        warning_faults = list(({'code': f['faultInst']['attributes']['code'],
                                'cause': f['faultInst']['attributes']['cause'],
                                'severity': f['faultInst']['attributes']['severity'],
                                'descr': f['faultInst']['attributes']['descr'],
                                'dn': f['faultInst']['attributes']['dn']}
                               for f in warning_faults))
        major_faults = apic.get('/api/class/faultInst.json?query-target-filter='
                                'eq(faultInst.severity,"major")').json()['imdata']
        major_faults = list(({'code': f['faultInst']['attributes']['code'],
                              'cause': f['faultInst']['attributes']['cause'],
                              'severity': f['faultInst']['attributes']['severity'],
                              'descr': f['faultInst']['attributes']['descr'],
                              'dn': f['faultInst']['attributes']['dn']}
                             for f in major_faults))
        critical_faults = apic.get('/api/class/faultInst.json?query-target-filter='
                                   'eq(faultInst.severity,"critical")').json()['imdata']
        critical_faults = list(({'code': f['faultInst']['attributes']['code'],
                                 'cause': f['faultInst']['attributes']['cause'],
                                 'severity': f['faultInst']['attributes']['severity'],
                                 'descr': f['faultInst']['attributes']['descr'],
                                 'dn': f['faultInst']['attributes']['dn']}
                                for f in critical_faults))
        fault_count = len(minor_faults) + len(warning_faults) + len(major_faults) + len(critical_faults)
        # Collect all nodes (fabricNode)
        fabric_nodes = apic.collect_nodes()
        fabric_nodes = list((node['fabricNode']['attributes']['dn'] for node in fabric_nodes))

        snapshot = {
            'environment': apic.env.Name,
            'timestamp': timestamp.ctime(),
            'ip_routes': {
                'count': len(ip_routes),
                'ip_routes': ip_routes
            },
            'ospf_routes': {
                'count': len(ospf_routes),
                'ospf_routes': ospf_routes
            },
            'ospf_neighbors': {
                'count': len(ospf_neighbors),
                'ospf_neighbors': []
            },
            'bgp_peers': {
                'count': len(bgp_peers),
                'bgp_peers': bgp_peers
            },
            'faults': {
                'count': fault_count,
                'minor_faults': {
                    'count': len(minor_faults),
                    'minor_faults': minor_faults
                },
                'warning_faults': {
                    'count': len(warning_faults),
                    'warning_faults': warning_faults
                },
                'major_faults': {
                    'count': len(major_faults),
                    'major_faults': major_faults
                },
                'critical_faults': {
                    'count': len(critical_faults),
                    'critical_faults': critical_faults
                }
            },
            'endpoints': {
                'count': len(endpoints),
                'endpoints': endpoints
            },
            'fabric_nodes': {
                'count': len(fabric_nodes),
                'fabric_nodes': fabric_nodes
            }
        }

        # snapshot = JSONObject(snapshot)

        if f'{apic.env.Name}_status.json' in api.list_dir('pyapis/aci/status'):
            old_snapshot = api.get_file_content(f'pyapis/aci/status/{apic.env.Name}_status.json')
            old_snapshot = json.loads(old_snapshot)

            comparison = {
                'environment': apic.env.Name,
                'previous_status_datetime': old_snapshot['timestamp'],
                'current_status_datetime': snapshot['timestamp'],
                'ip_route_diff': snapshot['ip_routes']['count'] - old_snapshot['ip_routes']['count'],
                'ospf_route_diff': snapshot['ospf_routes']['count'] - old_snapshot['ospf_routes']['count'],
                'ospf_neighbor_diff': snapshot['ospf_neighbors']['count'] - old_snapshot['ospf_neighbors']['count'],
                'bgp_peers_diff': snapshot['bgp_peers']['count'] - old_snapshot['bgp_peers']['count'],
                'endpoints_diff': snapshot['endpoints']['count'] - old_snapshot['endpoints']['count'],
                'fabric_nodes_diff': snapshot['fabric_nodes']['count'] - old_snapshot['fabric_nodes']['count'],
                'faults_diff': snapshot['faults']['count'] - old_snapshot['faults']['count'],
                'minor_faults_diff': int(snapshot['faults']['minor_faults']['count']) -
                int(old_snapshot['faults']['minor_faults']['count']),
                'warning_faults_diff': int(snapshot['faults']['warning_faults']['count']) -
                int(old_snapshot['faults']['warning_faults']['count']),
                'major_faults_diff': int(snapshot['faults']['major_faults']['count']) -
                int(old_snapshot['faults']['major_faults']['count']),
                'critical_faults_diff': int(snapshot['faults']['critical_faults']['count']) -
                int(old_snapshot['faults']['critical_faults']['count']),
                'endpoints': [],
                'ip_routes': [],
                'ospf_routes': [],
                'ospf_neighbors': [],
                'bgp_peers': [],
                'perf_counter': str()
            }

            # Find missing endpoints
            for endpoint in old_snapshot['endpoints']['endpoints']:
                if endpoint not in snapshot['endpoints']['endpoints']:
                    comparison['endpoints'].append(endpoint)
            # Find missing routes
            for route in old_snapshot['ip_routes']['ip_routes']:
                if route not in snapshot['ip_routes']['ip_routes']:
                    comparison['ip_routes'].append(route)
            # Find missing ospf routes
            for route in old_snapshot['ospf_routes']['ospf_routes']:
                if route not in snapshot['ospf_routes']['ospf_routes']:
                    comparison['ospf_routes'].append(route)
            # Find missing Neighbors
            comparison['ospf_neighbors'] = []
            for neighbor in old_snapshot['ospf_neighbors']['ospf_neighbors']:
                if neighbor not in snapshot['ospf_neighbors']['ospf_neighbors']:
                    comparison['ospf_neighbors'].append(neighbor)
            # Find missing bgp peers
            comparison['bgp_peers'] = []
            for peer in old_snapshot['bgp_peers']['bgp_peers']:
                if peer not in snapshot['bgp_peers']['bgp_peers']:
                    comparison['bgp_peers'].append(peer)
            # Find new faults
            # if comparison['minor_faults_diff'] > 0:
            #     comparison['minor_faults'] = []
            #     for fault in snapshot.faults.minor_faults.minor_faults:
            #         if fault.dict() not in list(f.dict() for f in old_snapshot.faults.minor_faults.minor_faults):
            #             comparison['minor_faults'].append(fault.dict())
            # if comparison['warning_faults_diff'] > 0:
            #     comparison['warning_faults'] = []
            #     for fault in snapshot.faults.warning_faults.warning_faults:
            #         if fault.dict() not in list(f.dict() for f in old_snapshot.faults.warning_faults.warning_faults):
            #             comparison['warning_faults'].append(fault.dict())
            # if comparison['major_faults_diff'] > 0:
            #     comparison['major_faults'] = []
            #     for fault in snapshot.faults.major_faults.major_faults:
            #         if fault.dict() not in list(f.dict() for f in old_snapshot.faults.major_faults.major_faults):
            #             comparison['major_faults'].append(fault.dict())
            # if comparison['critical_faults_diff'] > 0:
            #     comparison['critical_faults'] = []
            #     for fault in snapshot.faults.critical_faults.critical_faults:
            #         if fault.dict() not in list(f.dict() for f in old_snapshot.faults.critical_faults.
            #         critical_faults):
            #             comparison['critical_faults'].append(fault.dict())

            comparison['perf_counter'] = f'Ran for {round(time.perf_counter() - start)} seconds'
            api.update_file(file_path=f'pyapis/aci/status/{apic.env.Name}_status.json', message='fabric_status',
                            content=json.dumps(snapshot))
            return 200, comparison
        else:
            api.add_file(file_path=f'pyapis/aci/status/{apic.env.Name}_status.json', message='fabric_status',
                         content=json.dumps(snapshot))
            del api
            return 200, {'message': 'Initial status report created'}
